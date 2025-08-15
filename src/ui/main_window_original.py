"""
Main Window for DICOM Viewer
MVP Phase 1: Basic viewer with frame navigation
"""

import logging

logger = logging.getLogger(__name__)

import os
from pathlib import Path
from typing import Optional

import numpy as np
from PyQt6.QtCore import QSettings, Qt, QTimer
from PyQt6.QtGui import QAction, QFont, QKeySequence
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSlider,
    QStatusBar,
    QVBoxLayout,
    QWidget,
    QProgressDialog,
    QStackedWidget,
    QSizePolicy,
)

from ..analysis.qca_analysis import QCAAnalysis as QCAAnalyzer
from ..core.dicom_parser import DicomParser
from ..core.dicom_folder_loader import DicomFolderLoader, DicomSeries
from ..core.ekg_parser import EKGParser
from .activity_bar import ActivityBar
from .calibration_angiopy_widget import CalibrationAngioPyWidget
from .ekg_viewer_widget import EKGViewerWidget
from .enhanced_viewer_widget import EnhancedDicomViewer
from .projection_selection_dialog import ProjectionSelectionDialog
from .qca_widget import QCAWidget
from .segmentation_widget import SegmentationWidget


class MainWindow(QMainWindow):
    """Main application window"""

    def __init__(self):
        super().__init__()
        self.dicom_parser = DicomParser()
        self.ekg_parser = EKGParser()
        self.qca_analyzer = QCAAnalyzer()
        # Enable advanced diameter measurement for better accuracy
        self.qca_analyzer.use_advanced_diameter = False

        # Global reference diameter (calculated after sequential processing)
        self.global_reference_diameter = None

        self.current_file: Optional[Path] = None
        self.calibration_factor: Optional[float] = None  # User calibration
        self.fallback_calibration: Optional[float] = None  # From DICOM metadata
        self.is_playing = False
        self.play_timer = QTimer()
        self.play_timer.timeout.connect(self.next_frame_auto)
        self.frame_timestamps: Optional[np.ndarray] = None

        # Timer for track button blinking
        self.track_button_timer = QTimer()
        self.track_button_timer.timeout.connect(self.toggle_track_button_style)
        self.track_button_blink_state = False

        # Projection management
        self.all_projections = []  # List of all available projections
        self.current_projection_index = 0
        self.current_dicomdir_path = None

        # Analysis widgets and dialogs
        self.qca_widget: Optional[QCAWidget] = None
        self.segmentation_widget: Optional[SegmentationWidget] = None

        # Analysis dialogs removed - using accordion panels instead

        # Initialize AngioPy segmentation model using singleton
        from ..core.model_manager import ModelManager

        self.model_manager = ModelManager.instance()
        self.segmentation_model = self.model_manager.get_segmentation_model(auto_download=True)

        self.init_ui()
        self.load_settings()

    def init_ui(self):
        """Initialize the UI"""
        self.setWindowTitle("Coronary Clear Vision - DICOM Viewer")

        # Get available screen geometry
        from PyQt6.QtWidgets import QApplication

        # Get primary screen
        screen = QApplication.primaryScreen()
        if screen:
            available_rect = screen.availableGeometry()
            # Use 95% of available screen size for better fit
            width = min(int(available_rect.width() * 0.95), 1800)
            height = min(int(available_rect.height() * 0.95), 1100)

            # Center the window
            x = (available_rect.width() - width) // 2
            y = (available_rect.height() - height) // 2

            self.setGeometry(x, y, width, height)
        else:
            # Fallback for HD 720p displays
            self.setGeometry(100, 50, 1440, 810)

        # Set minimum size for HD 720p displays
        self.setMinimumSize(1280, 720)

        # Wayland compatibility: avoid maximize issues
        self.setWindowState(Qt.WindowState.WindowNoState)

        # Set application-wide font (optimized for HD displays)
        from PyQt6.QtGui import QFont

        app_font = QFont()
        app_font.setPointSize(11)  # Slightly larger for HD displays
        self.setFont(app_font)

        # Set application style
        self.setStyleSheet(
            """
            QTabBar::tab {
                background-color: #E3F2FD;
                color: #1976D2;
                padding: 8px 16px;
                margin-right: 2px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #1976D2;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #BBDEFB;
            }
        """
        )

        # Set focus policy to ensure keyboard events are captured
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main vertical layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Main horizontal layout with two sections
        main_h_widget = QWidget()
        main_h_layout = QHBoxLayout(main_h_widget)
        main_h_layout.setContentsMargins(0, 0, 0, 0)
        main_h_layout.setSpacing(5)

        # First create viewer widget (needed by accordion panel)
        self.viewer_widget = EnhancedDicomViewer()
        # Disable tracking by default
        self.viewer_widget.set_tracking_enabled(False)
        # Set main window reference
        self.viewer_widget.main_window = self

        # Connect frame range selection signal
        self.viewer_widget.frame_range_selected.connect(self.on_frame_range_selected)

        # Left side: Activity bar + Content panel
        left_container = QWidget()
        left_container.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Preferred)
        left_layout = QHBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)

        # Activity bar (VSCode style)
        self.activity_bar = ActivityBar()
        self.activity_bar.mode_changed.connect(self.on_activity_mode_changed)
        left_layout.addWidget(self.activity_bar)

        # Content panel (stacked widget)
        self.content_panel = self.create_content_panel()
        # Make content panel responsive - size to content
        self.content_panel.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        left_layout.addWidget(self.content_panel)

        # Set stretch factors - activity bar fixed, content panel can grow
        left_layout.setStretchFactor(self.activity_bar, 0)
        left_layout.setStretchFactor(self.content_panel, 1)

        main_h_layout.addWidget(left_container)

        # Right: Viewer and EKG stacked vertically
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 5, 5, 5)  # No left margin for right alignment
        right_layout.setSpacing(5)
        right_layout.setAlignment(Qt.AlignmentFlag.AlignRight)  # Align content to right

        # DICOM viewer with timeline
        viewer_widget = QWidget()
        viewer_layout = QVBoxLayout(viewer_widget)
        viewer_layout.setContentsMargins(0, 0, 0, 0)
        viewer_layout.setSpacing(0)

        # Add DICOM viewer
        viewer_layout.addWidget(self.viewer_widget)

        # Timeline controls
        nav_widget = self.create_navigation_widget()
        nav_widget.setMaximumHeight(60)  # Reduced height
        viewer_layout.addWidget(nav_widget)

        viewer_layout.setStretchFactor(self.viewer_widget, 1)
        right_layout.addWidget(viewer_widget)

        # EKG viewer
        self.ekg_viewer = EKGViewerWidget()
        self.ekg_viewer.setMaximumHeight(250)  # Increased height for better visibility
        self.ekg_viewer.setMinimumHeight(200)  # Increased minimum height
        right_layout.addWidget(self.ekg_viewer)

        # Set stretch factors for vertical layout
        right_layout.setStretchFactor(viewer_widget, 3)
        right_layout.setStretchFactor(self.ekg_viewer, 1)

        main_h_layout.addWidget(right_widget)

        # Set stretch factors for horizontal layout
        # Left panel gets 0 (fixed width), right side gets all remaining space
        main_h_layout.setStretchFactor(left_container, 0)
        main_h_layout.setStretchFactor(right_widget, 1)

        main_layout.addWidget(main_h_widget)

        # Create menus
        self.create_menus()

        # Create status bar
        self.status_bar = QStatusBar()
        self.status_bar.setStyleSheet("QStatusBar { font-size: 11px; }")  # Reduced font size
        self.setStatusBar(self.status_bar)
        self.update_status("")

        # Preload AngioPy model asynchronously after UI is ready
        from ..core.model_manager import ModelManager

        model_manager = ModelManager.instance()

        def model_progress(status: str, percent: int):
            if percent >= 0:
                self.update_status(f"{status} ({percent}%)")
            else:
                self.update_status(status)

        model_manager.preload_model_async(auto_download=True, progress_callback=model_progress)
        self.update_status("Loading AngioPy model in background...")

        # Connect signals
        self.viewer_widget.zoom_changed.connect(self.update_zoom_status)
        self.viewer_widget.pixel_info_changed.connect(self.update_pixel_info)
        self.viewer_widget.points_changed.connect(self.update_tracking_buttons)
        self.ekg_viewer.time_clicked.connect(self.on_ekg_time_clicked)

        # Connect segmentation signal (will be used when segmentation mode is active)
        self.viewer_widget.segmentation_point_clicked.connect(
            self.on_viewer_segmentation_point_clicked
        )

        # Connect calibration signal
        self.viewer_widget.calibration_point_clicked.connect(
            self.on_viewer_calibration_point_clicked
        )

    def create_navigation_widget(self) -> QWidget:
        """Create frame navigation controls"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 0, 0, 0)  # Small left margin

        # Add stretch to push controls to the right
        layout.addStretch(1)

        # Frame slider - fixed width
        self.frame_slider = QSlider(Qt.Orientation.Horizontal)
        self.frame_slider.setMinimum(0)
        self.frame_slider.setMaximum(0)
        self.frame_slider.setMinimumHeight(25)  # Reduced height for compact view
        self.frame_slider.setFixedWidth(160)  # Reduced by 20% (200 * 0.8)
        self.frame_slider.valueChanged.connect(self.on_frame_changed)
        self.frame_slider.setEnabled(False)
        self.frame_slider.setStyleSheet(
            """
            QSlider::groove:horizontal {
                height: 10px;
                background: #d3d3d3;
                border-radius: 5px;
            }
            QSlider::handle:horizontal {
                background: #5599ff;
                width: 20px;
                height: 20px;
                border-radius: 10px;
                margin: -5px 0;
            }
        """
        )
        layout.addWidget(self.frame_slider)

        # Frame label
        self.frame_label = QLabel("Frame: 0/0")
        self.frame_label.setMinimumWidth(96)  # Reduced by 20% (120 * 0.8)
        self.frame_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.frame_label.setStyleSheet("QLabel { font-size: 12px; font-weight: bold; }")
        layout.addWidget(self.frame_label)

        # Right side - All controls in compact group
        # Previous projection button
        self.prev_projection_button = QPushButton("⏮")
        self.prev_projection_button.setFixedSize(32, 32)  # Compact size
        self.prev_projection_button.setToolTip("Previous Projection")
        self.prev_projection_button.clicked.connect(self.previous_projection)
        self.prev_projection_button.setEnabled(False)
        self.prev_projection_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.prev_projection_button)

        # Previous frame button
        self.prev_frame_button = QPushButton("◀")
        self.prev_frame_button.setFixedSize(32, 32)  # Compact size
        self.prev_frame_button.setToolTip("Previous Frame")
        self.prev_frame_button.clicked.connect(self.previous_frame)
        self.prev_frame_button.setEnabled(False)
        self.prev_frame_button.setStyleSheet(
            """
            QPushButton {
                font-size: 18px;
                font-weight: bold;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.prev_frame_button)

        # Play/Pause button
        self.play_button = QPushButton("▶")
        self.play_button.setFixedSize(40, 32)  # Compact size
        self.play_button.setToolTip("Play/Pause")
        self.play_button.clicked.connect(self.toggle_play)
        self.play_button.setEnabled(False)
        self.play_button.setStyleSheet(
            """
            QPushButton {
                font-size: 14px;
                font-weight: bold;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.play_button)

        # Next frame button
        self.next_frame_button = QPushButton("▶")
        self.next_frame_button.setFixedSize(32, 32)  # Compact size
        self.next_frame_button.setToolTip("Next Frame")
        self.next_frame_button.clicked.connect(self.next_frame)
        self.next_frame_button.setEnabled(False)
        self.next_frame_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.next_frame_button)

        # Next projection button
        self.next_projection_button = QPushButton("⏭")
        self.next_projection_button.setFixedSize(32, 32)  # Compact size
        self.next_projection_button.setToolTip("Next Projection")
        self.next_projection_button.clicked.connect(self.next_projection)
        self.next_projection_button.setEnabled(False)
        self.next_projection_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.next_projection_button)

        # Projection selection button
        self.projection_select_button = QPushButton("📁")
        self.projection_select_button.setFixedSize(32, 32)  # Compact size
        self.projection_select_button.setToolTip("Select Projection")
        self.projection_select_button.clicked.connect(self.show_projection_dialog)
        self.projection_select_button.setEnabled(False)
        self.projection_select_button.setStyleSheet(
            """
            QPushButton {
                font-size: 14px;
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #999999;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999999;
                border: 1px solid #dddddd;
            }
        """
        )
        layout.addWidget(self.projection_select_button)

        # Add separator
        layout.addSpacing(10)  # Minimal spacing

        # Tracking controls
        # Track backward button
        self.track_backward_button = QPushButton("<-T")
        self.track_backward_button.setFixedSize(40, 40)  # Reduced size
        self.track_backward_button.setToolTip("Track Points to Previous Frame")
        self.track_backward_button.clicked.connect(self.track_backward)
        self.track_backward_button.setEnabled(False)
        self.track_backward_button.setStyleSheet(
            "QPushButton { font-size: 12px; font-weight: bold; }"
        )
        layout.addWidget(self.track_backward_button)

        # Track label
        track_label = QLabel("Track")
        track_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #673AB7;")
        track_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(track_label)

        # Track forward button
        self.track_forward_button = QPushButton("T->")
        self.track_forward_button.setFixedSize(40, 40)  # Reduced size
        self.track_forward_button.setToolTip("Track Points to Next Frame")
        self.track_forward_button.clicked.connect(self.track_forward)
        self.track_forward_button.setEnabled(False)
        self.track_forward_button.setStyleSheet(
            "QPushButton { font-size: 12px; font-weight: bold; }"
        )
        layout.addWidget(self.track_forward_button)

        # Track All button moved to left panel

        return widget

    def create_content_panel(self) -> QStackedWidget:
        """Create VSCode-style content panel with stacked widgets"""
        stack = QStackedWidget()
        stack.setStyleSheet(
            """
            QStackedWidget {
                background-color: #f5f5f5;
                border: 1px solid #e0e0e0;
            }
            QWidget {
                background-color: #f5f5f5;
                color: #333333;
            }
            QLabel {
                color: #333333;
                padding: 5px;
            }
            QPushButton {
                background-color: #0e639c;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 3px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #1177bb;
            }
            QPushButton:checked {
                background-color: #094771;
            }
            QGroupBox {
                color: #333333;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QCheckBox {
                color: #333333;
            }
        """
        )

        # Initialize analysis widgets
        self._init_analysis_widgets()

        # Store panel indices
        self.panel_indices = {}

        # 1. Calibration panel
        calibration_container = QWidget()
        calibration_container.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred
        )
        calibration_layout = QVBoxLayout(calibration_container)
        calibration_layout.setContentsMargins(10, 10, 10, 10)
        calibration_layout.addWidget(
            QLabel("CALIBRATION", styleSheet="font-weight: bold; font-size: 12px;")
        )
        calibration_layout.addWidget(self.calibration_widget)
        calibration_layout.addStretch()
        self.panel_indices["calibration"] = stack.addWidget(calibration_container)

        # 2. Tracking panel
        tracking_widget = QWidget()
        tracking_layout = QVBoxLayout(tracking_widget)
        tracking_layout.setContentsMargins(10, 10, 10, 10)

        tracking_layout.addWidget(
            QLabel("TRACKING", styleSheet="font-weight: bold; font-size: 12px;")
        )

        tracking_info = QLabel(
            "Click on the angiogram to add\ntracking points for vessels\n(unlimited points - 1, 2, 3, 4+ supported)"
        )
        tracking_info.setWordWrap(True)
        tracking_layout.addWidget(tracking_info)

        # Auto-activated when this mode is selected
        tracking_status = QLabel("Mode: Active")
        tracking_status.setStyleSheet("color: #4ec9b0; font-weight: bold;")
        tracking_layout.addWidget(tracking_status)

        # Add spacing before the analysis button
        tracking_layout.addSpacing(20)

        # Add Start Analysis button (previously Track All)
        self.track_all_button = QPushButton("🚀 Start Analysis")
        self.track_all_button.setMinimumHeight(40)
        self.track_all_button.setEnabled(False)
        self.track_all_button.clicked.connect(self.track_all_frames)
        self.track_all_button.setStyleSheet(
            """
            QPushButton { 
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
                background-color: #2E7D32;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:disabled {
                background-color: #B0BEC5;
                color: #ECEFF1;
            }
        """
        )
        self.track_all_button.setToolTip(
            "Track points through all frames and start sequential analysis"
        )
        tracking_layout.addWidget(self.track_all_button)

        # Removed centerline interpolation toggle - always use original length

        # Curvature-resistant centerline checkbox removed - always use default method

        # Light mask limiting checkbox removed - always use moderate/tight limiting

        # Add "Proceed to Analysis" button (hidden by default)
        self.proceed_to_analysis_btn = QPushButton("📊 Proceed to Analysis")
        self.proceed_to_analysis_btn.setMinimumHeight(35)
        self.proceed_to_analysis_btn.setStyleSheet(
            """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """
        )
        self.proceed_to_analysis_btn.clicked.connect(self.on_proceed_to_analysis)
        self.proceed_to_analysis_btn.setVisible(False)
        tracking_layout.addWidget(self.proceed_to_analysis_btn)

        tracking_layout.addStretch()
        self.panel_indices["tracking"] = stack.addWidget(tracking_widget)

        # 3. Segmentation panel
        segmentation_container = QWidget()
        segmentation_layout = QVBoxLayout(segmentation_container)
        segmentation_layout.setContentsMargins(10, 10, 10, 10)
        segmentation_layout.addWidget(
            QLabel("SEGMENTATION", styleSheet="font-weight: bold; font-size: 12px;")
        )
        segmentation_layout.addWidget(self.segmentation_widget)
        segmentation_layout.addStretch()
        self.panel_indices["segmentation"] = stack.addWidget(segmentation_container)

        # 4. QCA Analysis panel
        qca_container = QWidget()
        qca_layout = QVBoxLayout(qca_container)
        qca_layout.setContentsMargins(5, 5, 5, 5)
        qca_layout.setSpacing(5)
        qca_layout.addWidget(
            QLabel("QCA ANALYSIS", styleSheet="font-weight: bold; font-size: 12px;")
        )
        qca_layout.addWidget(self.qca_widget, 1)  # Add stretch factor to fill available space
        self.panel_indices["qca"] = stack.addWidget(qca_container)

        # 5. Batch Processing panel
        batch_widget = QWidget()
        batch_layout = QVBoxLayout(batch_widget)
        batch_layout.setContentsMargins(10, 10, 10, 10)

        batch_layout.addWidget(
            QLabel("BATCH PROCESSING", styleSheet="font-weight: bold; font-size: 12px;")
        )

        seq_btn = QPushButton("Start Sequential Processing")
        seq_btn.setMinimumHeight(35)
        seq_btn.clicked.connect(self.start_sequential_processing)
        batch_layout.addWidget(seq_btn)

        batch_info = QLabel("Process multiple frames\nautomatically")
        batch_info.setWordWrap(True)
        batch_layout.addWidget(batch_info)

        batch_layout.addStretch()
        self.panel_indices["batch"] = stack.addWidget(batch_widget)

        # 7. Export panel
        export_widget = QWidget()
        export_layout = QVBoxLayout(export_widget)
        export_layout.setContentsMargins(10, 10, 10, 10)

        export_layout.addWidget(QLabel("EXPORT", styleSheet="font-weight: bold; font-size: 12px;"))

        # Export buttons
        export_pdf_btn = QPushButton("Export Report as PDF")
        export_pdf_btn.setMinimumHeight(30)
        export_pdf_btn.clicked.connect(lambda: self.export_analysis_report("pdf"))
        export_layout.addWidget(export_pdf_btn)

        export_xlsx_btn = QPushButton("Export Report as Excel (XLSX)")
        export_xlsx_btn.setMinimumHeight(30)
        export_xlsx_btn.setStyleSheet(
            """
            QPushButton { 
                background-color: #4CAF50; 
                color: white; 
                font-weight: bold;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover { 
                background-color: #45a049; 
            }
        """
        )
        export_xlsx_btn.clicked.connect(lambda: self.export_analysis_report("xlsx"))
        export_layout.addWidget(export_xlsx_btn)

        export_txt_btn = QPushButton("Export Report as Text")
        export_txt_btn.setMinimumHeight(30)
        export_txt_btn.clicked.connect(lambda: self.export_analysis_report("txt"))
        export_layout.addWidget(export_txt_btn)

        export_layout.addStretch()
        self.panel_indices["export"] = stack.addWidget(export_widget)

        # Set default panel (calibration)
        stack.setCurrentIndex(self.panel_indices["calibration"])
        self.current_mode = "calibration"

        # Schedule size update after widget is shown
        QTimer.singleShot(100, self.update_content_panel_size)

        return stack

    def _init_analysis_widgets(self):
        """Initialize analysis widgets for accordion panel"""
        # Create calibration widget
        self.calibration_widget = CalibrationAngioPyWidget()
        self.calibration_widget.set_main_window(self)
        self.calibration_widget.set_segmentation_model(self.segmentation_model)

        # Create segmentation widget
        self.segmentation_widget = SegmentationWidget()
        self.segmentation_widget.set_segmentation_model(self.segmentation_model)
        self.segmentation_widget.set_main_window(self)

        # Create QCA widget
        self.qca_widget = QCAWidget()
        self.qca_widget.qca_analyzer = self.qca_analyzer
        self.qca_widget.set_main_window(self)  # Set reference for overlay controls

        # Connect signals
        self._connect_widget_signals()

    def update_content_panel_size(self):
        """Update content panel size based on current widget's content"""
        current_widget = self.content_panel.currentWidget()
        if current_widget:
            # Get the size hint and minimum size hint
            size_hint = current_widget.sizeHint()
            min_size_hint = current_widget.minimumSizeHint()

            # Calculate the actual content width
            # Use layout's content width if available
            if hasattr(current_widget, "layout") and current_widget.layout():
                layout = current_widget.layout()
                content_width = layout.sizeHint().width()
                min_content_width = layout.minimumSize().width()
            else:
                content_width = size_hint.width()
                min_content_width = min_size_hint.width()

            # Use the larger of the two, with some padding
            final_min_width = (
                max(content_width, min_content_width) + 40
            )  # 20px padding on each side

            # Ensure reasonable bounds
            final_min_width = max(final_min_width, 350)  # At least 350px
            final_min_width = min(final_min_width, 600)  # At most 600px

            # Set the content panel size
            self.content_panel.setMinimumWidth(int(final_min_width))
            self.content_panel.setMaximumWidth(int(final_min_width * 1.2))  # Allow 20% growth

            # Force layout update
            self.content_panel.updateGeometry()
            if self.content_panel.parent():
                self.content_panel.parent().updateGeometry()

    def on_activity_mode_changed(self, mode: str):
        """Handle activity bar mode changes"""
        # Switch panel
        if mode in self.panel_indices:
            self.content_panel.setCurrentIndex(self.panel_indices[mode])
            # Update content panel size based on current widget
            self.update_content_panel_size()

        # Deactivate previous mode
        if self.current_mode and self.current_mode != mode:
            if self.current_mode == "calibration":
                # Deactivate calibration widget
                if hasattr(self.calibration_widget, "deactivate"):
                    self.calibration_widget.deactivate()
            elif self.current_mode == "tracking":
                self.viewer_widget.set_tracking_enabled(False)
            elif self.current_mode == "segmentation":
                if (
                    hasattr(self.segmentation_widget, "mode_button")
                    and self.segmentation_widget.mode_button.isChecked()
                ):
                    self.segmentation_widget.mode_button.setChecked(False)
                    self.segmentation_widget.toggle_mode()

        # Activate new mode
        if mode == "calibration":
            # Automatically enable calibration mode
            self.calibration_widget.toggle_mode()
            if (
                hasattr(self.calibration_widget, "mode_button")
                and not self.calibration_widget.mode_button.isChecked()
            ):
                self.calibration_widget.mode_button.setChecked(True)
                self.calibration_widget.toggle_mode()
        elif mode == "tracking":
            self.viewer_widget.set_tracking_enabled(True)
        elif mode == "segmentation":
            if (
                hasattr(self.segmentation_widget, "mode_button")
                and not self.segmentation_widget.mode_button.isChecked()
            ):
                self.segmentation_widget.mode_button.setChecked(True)
                self.segmentation_widget.toggle_mode()
            # Set frame range
            total_frames = self.dicom_parser.get_frame_count()
            self.segmentation_widget.set_frame_range(total_frames)
        elif mode == "qca":
            # Pass calibration to QCA widget if available
            if self.calibration_factor:
                self.qca_widget.set_calibration(
                    self.calibration_factor, getattr(self, "calibration_details", None)
                )

            # Display current frame's QCA results if available
            current_frame = (
                self.viewer_widget.current_frame_index
                if hasattr(self.viewer_widget, "current_frame_index")
                else 0
            )
            if (
                hasattr(self.viewer_widget, "frame_qca_results")
                and current_frame in self.viewer_widget.frame_qca_results
            ):
                qca_result = self.viewer_widget.frame_qca_results[current_frame]
                if qca_result and qca_result.get("success"):
                    self.qca_widget.display_frame_results(current_frame, qca_result)

        # Update current mode
        self.current_mode = mode

        # Keep viewer focused for navigation
        self.viewer_widget.setFocus()

    def _connect_widget_signals(self):
        """Connect signals from analysis widgets"""
        # Calibration widget signals (CalibrationAngioPyWidget)
        self.calibration_widget.calibration_mode_changed.connect(self.on_calibration_mode_changed)
        self.calibration_widget.calibration_completed.connect(
            lambda factor, details: self.on_calibration_completed_widget(factor, details)
        )
        self.calibration_widget.overlay_settings_changed.connect(
            self.on_calibration_overlay_changed
        )

        # Segmentation widget signals
        self.segmentation_widget.segmentation_mode_changed.connect(
            self.on_segmentation_mode_changed
        )
        self.segmentation_widget.segmentation_completed.connect(self.update_segmentation_overlay)
        self.segmentation_widget.overlay_settings_changed.connect(
            self.update_segmentation_overlay_settings
        )
        self.segmentation_widget.qca_analysis_requested.connect(self.start_qca_from_segmentation)

        # QCA widget signals
        self.qca_widget.qca_started.connect(lambda: None)  # Not currently used
        self.qca_widget.qca_completed.connect(self.update_qca_overlay)
        self.qca_widget.overlay_changed.connect(
            lambda enabled, settings: self.update_qca_overlay_settings(settings)
        )
        self.qca_widget.calibration_requested.connect(self.start_calibration)

    def open_dicom(self):
        """Open DICOM file dialog"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open DICOM File",
            str(Path.home()),
            "All Files (*);;DICOM Files (*.dcm *.DCM);;DICOMDIR Files (DICOMDIR)",
        )

        if file_path:
            self.load_dicom_file(file_path)

    def restart_with_dicom(self, dicom_path: str, is_folder: bool = False):
        """DEPRECATED: This method is no longer used. 
        DICOM loading now happens without restarting the application.
        Kept for backward compatibility."""
        logger.warning("restart_with_dicom called but is deprecated. Using direct loading instead.")
        
        # Instead of restarting, clear and reload
        self.clear_all_analysis()
        
        if is_folder:
            self.load_dicom_folder_direct(dicom_path)
        else:
            self.load_dicom_file(dicom_path)

    def load_dicom_folder_direct(self, folder_path: str):
        """Load DICOM folder directly without dialog"""
        if folder_path and Path(folder_path).exists():
            # Check if we already have a DICOM loaded
            if hasattr(self, "dicom_parser") and self.dicom_parser is not None:
                # Ask user about loading new DICOM
                reply = QMessageBox.question(
                    self,
                    "Load New DICOM",
                    "Loading a new DICOM series will clear current analysis.\n\n"
                    "Do you want to continue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes,
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Clear all current analysis
                    self.clear_all_analysis()
                else:
                    return

            # Save for auto-load on next startup
            self.last_loaded_path = folder_path
            self.last_loaded_type = "folder"

            # Use new DicomFolderLoader
            folder_loader = DicomFolderLoader()

            # Show progress dialog
            progress = QProgressDialog("Loading DICOM folder...", None, 0, 0, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.show()

            try:
                if folder_loader.load_folder(folder_path):
                    series_list = folder_loader.get_series_list()
                    study_info = folder_loader.get_study_info()

                    if not series_list:
                        QMessageBox.warning(
                            self, "No Series Found", "No DICOM series found in the selected folder."
                        )
                        return

                    # Store folder loader for later use
                    self.folder_loader = folder_loader
                    self.all_series = series_list

                    if len(series_list) == 1:
                        # Single series - load directly
                        self.current_series_index = 0
                        self._load_series(series_list[0])

                        # Disable projection controls
                        self.prev_projection_button.setEnabled(False)
                        self.next_projection_button.setEnabled(False)
                        self.projection_select_button.setEnabled(False)
                    else:
                        # Multiple series - show selection dialog
                        from .dicom_projection_dialog import ProjectionSelectionDialog

                        dlg = ProjectionSelectionDialog(series_list, study_info, self)
                        if dlg.exec():
                            selected_series = dlg.selected_series
                            if selected_series:
                                # Find index of selected series
                                for idx, series in enumerate(series_list):
                                    if series.series_uid == selected_series.series_uid:
                                        self.current_series_index = idx
                                        break

                                self._load_series(selected_series)

                                # Enable projection controls if multiple series
                                self.prev_projection_button.setEnabled(
                                    self.current_series_index > 0
                                )
                                self.next_projection_button.setEnabled(
                                    self.current_series_index < len(self.all_series) - 1
                                )
                                self.projection_select_button.setEnabled(True)
                else:
                    QMessageBox.critical(
                        self, "Error", f"Failed to load DICOM folder:\n{folder_path}"
                    )
            finally:
                progress.close()

    def open_dicom_folder(self):
        """Open DICOM folder dialog"""
        folder_path = QFileDialog.getExistingDirectory(
            self, "Select DICOM Folder", str(Path.home()), QFileDialog.Option.ShowDirsOnly
        )

        if folder_path:
            self.load_dicom_folder_direct(folder_path)

    def open_default_dicom_folder(self):
        """Open the default DICOM folder"""
        default_folder = "/Users/fatihkoksal/Projelerim/Coronary_Clear_Vision/LOCAL"

        # Check if the folder exists
        if not Path(default_folder).exists():
            QMessageBox.warning(
                self,
                "Folder Not Found",
                f"Default DICOM folder not found:\n{default_folder}\n\n"
                "Please make sure the drive is mounted.",
            )
            return

        # Load the folder directly
        self.load_dicom_folder_direct(default_folder)

    def load_dicom_file(self, file_path: str):
        """Load a DICOM file"""
        # Check if we already have a DICOM loaded
        if hasattr(self, "dicom_parser") and self.dicom_parser is not None:
            # Ask user about loading new DICOM
            reply = QMessageBox.question(
                self,
                "Load New DICOM",
                "Loading a new DICOM file will clear current analysis.\n\n"
                "Do you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )

            if reply == QMessageBox.StandardButton.Yes:
                # Clear all current analysis
                self.clear_all_analysis()
            else:
                return

        # Save for auto-load on next startup
        self.last_loaded_path = file_path
        self.last_loaded_type = "file"

        # Check if it's a DICOMDIR - show projection selection dialog
        if Path(file_path).name == "DICOMDIR":
            projections = DicomParser.get_dicomdir_projections(file_path)

            if projections:
                # Store all projections and DICOMDIR path
                self.all_projections = projections
                self.current_dicomdir_path = file_path
                self.current_projection_index = 0

                # Enable projection controls
                self.prev_projection_button.setEnabled(len(projections) > 1)
                self.next_projection_button.setEnabled(len(projections) > 1)
                self.projection_select_button.setEnabled(True)

                # Show projection selection dialog
                dialog = ProjectionSelectionDialog(projections, self)
                if dialog.exec() == dialog.DialogCode.Accepted:
                    selected_file = dialog.get_selected_file_path()
                    if selected_file:
                        # Find index of selected projection
                        for i, proj in enumerate(self.all_projections):
                            if proj["file_path"] == selected_file:
                                self.current_projection_index = i
                                break
                        self._load_single_dicom(selected_file)
                return
            else:
                QMessageBox.warning(self, "Warning", "No projections found in DICOMDIR")
                return

        # Load single DICOM file
        self.all_projections = []  # Clear projections list
        self.current_dicomdir_path = None
        self.prev_projection_button.setEnabled(False)
        self.next_projection_button.setEnabled(False)
        self.projection_select_button.setEnabled(False)
        self._load_single_dicom(file_path)

    def clear_all_analysis(self):
        """Clear all analysis results and overlays"""
        try:
            # Reset EKG parser and viewer
            if hasattr(self, "ekg_parser") and self.ekg_parser:
                self.ekg_parser.reset()
            if hasattr(self, "ekg_viewer") and self.ekg_viewer:
                self.ekg_viewer.reset()

            # Clear all overlays from viewer
            if hasattr(self, "viewer_widget") and self.viewer_widget:
                try:
                    self.viewer_widget.clear_overlays()
                    self.viewer_widget.set_tracking_enabled(False)
                    self.viewer_widget.clear_user_points()
                except Exception as e:
                    logger.warning(f"Error clearing viewer overlays: {e}")

            # Clear QCA results
            if hasattr(self, "qca_widget") and self.qca_widget:
                try:
                    self.qca_widget.clear_results()
                except Exception as e:
                    logger.warning(f"Error clearing QCA results: {e}")

            # Clear RWS results
            if hasattr(self, "rws_widget") and self.rws_widget:
                try:
                    self.rws_widget.clear_analysis()
                except Exception as e:
                    logger.warning(f"Error clearing RWS results: {e}")

            if hasattr(self, "rws_enhanced_widget") and self.rws_enhanced_widget:
                try:
                    self.rws_enhanced_widget.clear_analysis()
                except Exception as e:
                    logger.warning(f"Error clearing enhanced RWS results: {e}")

            # Clear segmentation results - be careful here as it may trigger display_frame
            if hasattr(self, "segmentation_widget") and self.segmentation_widget:
                try:
                    # Instead of clear_mask which triggers display_frame, directly clear the mask
                    if hasattr(self.segmentation_widget, "current_mask"):
                        self.segmentation_widget.current_mask = None
                    if hasattr(self.segmentation_widget, "mask_history"):
                        self.segmentation_widget.mask_history.clear()
                except Exception as e:
                    logger.warning(f"Error clearing segmentation: {e}")

            # Reset batch segmentation results
            if hasattr(self, "batch_segmentation_results"):
                self.batch_segmentation_results = {}

            # Reset tracked range
            if hasattr(self, "tracked_range"):
                delattr(self, "tracked_range")

            # Hide proceed to analysis button if visible
            if hasattr(self, "proceed_to_analysis_btn"):
                self.proceed_to_analysis_btn.setVisible(False)

            # Clear calibration
            if hasattr(self, "calibration_factor"):
                self.calibration_factor = None

            # Reset cardiac phases
            if hasattr(self, "viewer_widget") and self.viewer_widget:
                self.viewer_widget.cardiac_phases = None
                self.viewer_widget.frame_timestamps = None

            logger.info("All analysis results and overlays cleared")

        except Exception as e:
            logger.error(f"Error in clear_all_analysis: {e}")

    def _load_series(self, series: DicomSeries):
        """Load a DICOM series"""
        if not hasattr(self, "folder_loader") or not self.folder_loader:
            logger.error("No folder loader available")
            return

        # Clear all previous analysis
        self.clear_all_analysis()

        dicom_data = self.folder_loader.load_series(series)
        if dicom_data:
            # Create a temporary file path for compatibility with existing code
            file_path = str(series.files[0]) if series.files else ""

            # Load using existing parser
            if self.dicom_parser.dicom_data is None:
                self.dicom_parser = DicomParser()

            self.dicom_parser.dicom_data = dicom_data

            # Try to access pixel data with error handling
            try:
                self.dicom_parser.pixel_array = dicom_data.pixel_array
            except AttributeError as e:
                if "pixel data" in str(e).lower():
                    QMessageBox.critical(
                        self,
                        "Invalid DICOM File",
                        f"Cannot load series: The DICOM file does not contain pixel data.\n\n"
                        f"Series: {series.series_description}\n"
                        f"This may be a non-image DICOM file (e.g., structured report, presentation state).",
                    )
                    return
                else:
                    raise
            except Exception as e:
                QMessageBox.critical(
                    self, "DICOM Loading Error", f"Failed to load pixel data: {str(e)}"
                )
                return

            self.dicom_parser.is_multi_frame = hasattr(dicom_data, "NumberOfFrames")
            self.dicom_parser.num_frames = (
                int(dicom_data.NumberOfFrames) if self.dicom_parser.is_multi_frame else 1
            )
            self.dicom_parser._extract_metadata()

            # Update status
            info = f"Series {series.series_number}: {series.series_description}"
            if series.primary_angle is not None and series.secondary_angle is not None:
                info += f" | LAO/RAO: {series.primary_angle:.1f}° | CRAN/CAUD: {series.secondary_angle:.1f}°"
            self.status_bar.showMessage(f"Loaded: {info}")

            # Continue with normal loading process
            self._load_single_dicom(file_path)

    def _load_single_dicom(self, file_path: str):
        """Load a single DICOM file"""
        # Clear all previous analysis
        self.clear_all_analysis()

        if self.dicom_parser.load_dicom(file_path):
            self.current_file = Path(file_path)
            self.add_to_recent(file_path)

            # Set fallback calibration from DICOM metadata
            if self.dicom_parser.pixel_spacing is not None:
                self.fallback_calibration = self.dicom_parser.pixel_spacing

                # Don't automatically use DICOM pixel spacing - it's often unreliable
                # User should always perform manual catheter calibration
                if self.calibration_factor is None:
                    logger.warning(
                        f"DICOM pixel spacing available ({self.fallback_calibration:.5f} mm/pixel) but not using it automatically"
                    )
                    logger.warning(
                        "User should perform manual catheter calibration for accurate measurements"
                    )
                    self.update_status(
                        "⚠️ Please perform catheter calibration for accurate measurements"
                    )

                # Pass pixel spacing to calibration widget
                if hasattr(self, "calibration_widget"):
                    self.calibration_widget.set_dicom_pixel_spacing(self.fallback_calibration)

            # Update UI
            self.viewer_widget.set_dicom_parser(self.dicom_parser)

            # Update frame navigation
            num_frames = self.dicom_parser.num_frames
            self.frame_slider.setMaximum(num_frames - 1)
            self.frame_slider.setValue(0)
            self.frame_slider.setEnabled(num_frames > 1)
            self.play_button.setEnabled(num_frames > 1)
            self.prev_frame_button.setEnabled(num_frames > 1)
            self.next_frame_button.setEnabled(num_frames > 1)
            self.update_frame_label(0)

            # Update window/level menu
            self.update_wl_menu()

            # Update status
            self.update_status(f"Loaded: {self.current_file.name}")

            # Display first frame
            self.viewer_widget.display_frame(0)
            self.viewer_widget.fit_to_window()

            # Reset heartbeat overlay
            self.viewer_widget.reset_heartbeat_overlay()

            # Update segmentation widget - call on_frame_changed for first frame
            if hasattr(self, "segmentation_widget") and self.segmentation_widget:
                self.segmentation_widget.on_frame_changed(0)

            # Update tracking buttons
            self.update_tracking_buttons()

            # Try to extract EKG data
            self.extract_and_display_ekg()

            # Extract frame timestamps for synchronization
            self.extract_frame_timestamps()
            # Share timestamps with viewer widget
            if hasattr(self, "frame_timestamps"):
                self.viewer_widget.frame_timestamps = self.frame_timestamps

            # Check synchronization status
            self.check_ekg_sync_status()

            # Automatically enable calibration mode after loading DICOM
            self.activity_bar.set_active_mode("calibration")
            self.on_activity_mode_changed("calibration")

            # Start automatic playback for multi-frame DICOMs
            if self.dicom_parser.is_multi_frame and self.dicom_parser.num_frames > 1:
                # Small delay to ensure UI is fully updated
                QTimer.singleShot(100, self.start_auto_playback)
        else:
            QMessageBox.critical(self, "Error", "Failed to load DICOM file")

    def on_frame_changed(self, value: int):
        """Handle frame slider change"""
        if self.dicom_parser.pixel_array is not None:
            self.viewer_widget.display_frame(value)
            self.update_frame_label(value)
            # Update EKG marker for bidirectional sync
            self.update_ekg_frame_marker(value)
            # Update tracking button states
            self.update_tracking_buttons()

            # Reset heartbeat counter if we're at frame 0
            if value == 0:
                self.viewer_widget.reset_heartbeat_overlay()

            # Check for R-peaks at current frame time
            self.check_r_peak_at_frame(value)

            # Maintain calibration mode if active
            if (
                self.current_mode == "calibration"
                and hasattr(self.calibration_widget, "is_active")
                and self.calibration_widget.is_active
            ):
                # Re-enable calibration mode in viewer if it was disabled
                self.viewer_widget.set_calibration_mode(True)

            # Load stored QCA/segmentation results if available
            if (
                hasattr(self.viewer_widget, "frame_qca_results")
                and value in self.viewer_widget.frame_qca_results
            ):
                qca_result = self.viewer_widget.frame_qca_results[value]
                if qca_result and qca_result.get("success"):
                    settings = (
                        self.qca_widget.get_overlay_settings()
                        if hasattr(self, "qca_widget")
                        else {
                            "show_measurements": True,
                            "show_stenosis": True,
                            "show_reference": True,
                            "show_centerline": True,
                            "color": "Yellow",
                        }
                    )
                    # Add enabled flag
                    settings["enabled"] = True
                    self.viewer_widget.set_qca_overlay(qca_result, settings)

                    # Update QCA widget with frame-specific results if in QCA mode
                    if self.current_mode == "qca" and hasattr(self, "qca_widget"):
                        self.qca_widget.display_frame_results(value, qca_result)

            if (
                hasattr(self.viewer_widget, "frame_segmentation_results")
                and value in self.viewer_widget.frame_segmentation_results
            ):
                seg_result = self.viewer_widget.frame_segmentation_results[value]
                if seg_result and seg_result.get("success"):
                    settings = {
                        "show_overlay": True,
                        "opacity": 1.0,
                        "color": "Red",
                        "contour_only": False,
                    }
                    self.viewer_widget.set_segmentation_overlay(seg_result, settings)
            else:
                # Clear overlays if no results for this frame
                self.viewer_widget.overlay_item.segmentation_mask = None
                self.viewer_widget.overlay_item.qca_results = None
                self.viewer_widget._request_update()

            # Update segmentation widget - call on_frame_changed to clear points and import tracking points
            if hasattr(self, "segmentation_widget") and self.segmentation_widget:
                self.segmentation_widget.on_frame_changed(value)

    def update_frame_label(self, current_frame: int):
        """Update frame label"""
        total_frames = self.dicom_parser.num_frames
        self.frame_label.setText(f"Frame: {current_frame + 1}/{total_frames}")

    def start_auto_playback(self):
        """Start automatic playback after DICOM load"""
        if not self.is_playing and self.play_button.isEnabled():
            # Don't exit calibration mode for auto-playback
            # Users can still see the video while calibrating

            self.is_playing = True
            self.play_timer.start(67)  # ~15 fps (standard cardiac cine rate)
            self.play_button.setText("⏸")

            logger.info("Auto-playback started")

    def toggle_play(self):
        """Toggle play/pause"""
        if self.is_playing:
            self.is_playing = False
            self.play_timer.stop()
            self.play_button.setText("▶")
        else:
            # Exit calibration mode before starting playback
            if self.viewer_widget.interaction_mode == "calibrate":
                self.viewer_widget.stop_calibration()

            # Clear any calibration graphics during playback
            if hasattr(self.viewer_widget, "overlay_item"):
                self.viewer_widget.overlay_item.calibration_line = None
                self.viewer_widget.overlay_item.update()

            self.is_playing = True
            self.play_timer.start(67)  # ~15 fps (standard cardiac cine rate)
            self.play_button.setText("⏸")

    def next_frame(self):
        """Advance to next frame (manual)"""
        current = self.frame_slider.value()
        if current < self.frame_slider.maximum():
            self.frame_slider.setValue(current + 1)

    def next_frame_auto(self):
        """Advance to next frame (auto-play with loop)"""
        current = self.frame_slider.value()
        if current < self.frame_slider.maximum():
            self.frame_slider.setValue(current + 1)
        else:
            # Loop back to beginning of range
            self.frame_slider.setValue(self.frame_slider.minimum())
            # Reset heartbeat counter when looping
            self.viewer_widget.reset_heartbeat_overlay()

    def previous_frame(self):
        """Go to previous frame"""
        current = self.frame_slider.value()
        if current > 0:
            self.frame_slider.setValue(current - 1)

    def next_projection(self):
        """Load next projection"""
        logger.info("Loading next projection...")

        # Check if we have series from folder loader
        if hasattr(self, "all_series") and self.all_series and len(self.all_series) > 1:
            self.current_series_index = (self.current_series_index + 1) % len(self.all_series)
            self._load_series(self.all_series[self.current_series_index])
        # Fallback to old projection system
        elif self.all_projections and len(self.all_projections) > 1:
            self.current_projection_index = (self.current_projection_index + 1) % len(
                self.all_projections
            )
            self._load_projection_at_index(self.current_projection_index)

    def previous_projection(self):
        """Load previous projection"""
        logger.info("Loading previous projection...")

        # Check if we have series from folder loader
        if hasattr(self, "all_series") and self.all_series and len(self.all_series) > 1:
            self.current_series_index = (self.current_series_index - 1) % len(self.all_series)
            self._load_series(self.all_series[self.current_series_index])
        # Fallback to old projection system
        elif self.all_projections and len(self.all_projections) > 1:
            self.current_projection_index = (self.current_projection_index - 1) % len(
                self.all_projections
            )
            self._load_projection_at_index(self.current_projection_index)

    def _load_projection_at_index(self, index: int):
        """Load projection at given index"""
        if 0 <= index < len(self.all_projections):
            projection = self.all_projections[index]
            self._load_single_dicom(projection["file_path"])

    def show_projection_dialog(self):
        """Show projection selection dialog"""
        # Check if we have series from folder loader
        if hasattr(self, "all_series") and self.all_series and hasattr(self, "folder_loader"):
            from .dicom_projection_dialog import ProjectionSelectionDialog

            study_info = self.folder_loader.get_study_info()
            dialog = ProjectionSelectionDialog(self.all_series, study_info, self)
            if dialog.exec() == dialog.DialogCode.Accepted:
                selected_series = dialog.selected_series
                if selected_series:
                    # Find index of selected series
                    for i, series in enumerate(self.all_series):
                        if series.series_uid == selected_series.series_uid:
                            self.current_series_index = i
                            break
                    self._load_series(selected_series)
        # Fallback to old projection system
        elif self.all_projections:
            dialog = ProjectionSelectionDialog(self.all_projections, self)
            if dialog.exec() == dialog.DialogCode.Accepted:
                selected_file = dialog.get_selected_file_path()
                if selected_file:
                    # Find index of selected projection
                    for i, proj in enumerate(self.all_projections):
                        if proj["file_path"] == selected_file:
                            self.current_projection_index = i
                            break
                    self._load_single_dicom(selected_file)

    def update_status(self, message: str):
        """Update status bar"""
        if message:
            self.status_bar.showMessage(message)
        else:
            # Show default info
            if self.dicom_parser.pixel_array is not None:
                info = f"Size: {self.dicom_parser.metadata['columns']}x{self.dicom_parser.metadata['rows']}"
                if self.dicom_parser.num_frames > 1:
                    info += f" | Frames: {self.dicom_parser.num_frames}"
                self.status_bar.showMessage(info)

    def track_forward(self):
        """Track points to next frame"""
        if self.viewer_widget.track_points_to_next_frame():
            # Move to next frame after successful tracking
            self.next_frame()
            self.update_status("Points tracked to next frame")
        else:
            self.update_status("Failed to track points")

    def track_backward(self):
        """Track points to previous frame"""
        if self.viewer_widget.track_points_to_previous_frame():
            # Move to previous frame after successful tracking
            self.previous_frame()
            self.update_status("Points tracked to previous frame")
        else:
            self.update_status("Failed to track points")

    def update_tracking_buttons(self):
        """Update tracking button states and make track all button blink when ready"""
        if not hasattr(self, "viewer_widget") or not hasattr(self, "track_all_button"):
            return

        # Get current tracking points - check both user_points and frame_points
        current_points = []
        if hasattr(self.viewer_widget, "overlay_item"):
            # First check user_points (for multi-frame mode)
            if (
                hasattr(self.viewer_widget.overlay_item, "user_points")
                and self.viewer_widget.overlay_item.user_points
            ):
                current_points = self.viewer_widget.overlay_item.user_points
            # Also check frame-specific points
            elif (
                hasattr(self.viewer_widget.overlay_item, "frame_points")
                and self.viewer_widget.current_frame_index
                in self.viewer_widget.overlay_item.frame_points
            ):
                current_points = self.viewer_widget.overlay_item.frame_points[
                    self.viewer_widget.current_frame_index
                ]

        num_points = len(current_points) if current_points else 0
        has_points = num_points > 0

        # Can track forward if not at last frame and have points
        can_track_forward = (
            has_points
            and hasattr(self, "frame_slider")
            and self.frame_slider.value() < self.frame_slider.maximum()
        )
        if hasattr(self, "track_forward_button"):
            self.track_forward_button.setEnabled(can_track_forward)

        # Can track backward if not at first frame and have points
        can_track_backward = (
            has_points and hasattr(self, "frame_slider") and self.frame_slider.value() > 0
        )
        if hasattr(self, "track_backward_button"):
            self.track_backward_button.setEnabled(can_track_backward)

        # Start Analysis button logic with enhanced blinking
        if num_points >= 2:
            # Enable Start Analysis button
            self.track_all_button.setEnabled(True)

            # Start enhanced blinking in tracking mode
            if getattr(self, "current_mode", "") == "tracking":
                self.start_track_button_blinking()
                if not hasattr(self, "_tracking_ready_logged"):
                    logger.info(
                        f"Start Analysis button ready - {num_points} tracking points available"
                    )
                    self._tracking_ready_logged = True
        else:
            # Disable Start Analysis button and stop blinking
            self.track_all_button.setEnabled(False)
            self.stop_track_button_blinking()

            # Clear logging flag
            if hasattr(self, "_tracking_ready_logged"):
                delattr(self, "_tracking_ready_logged")

            if num_points == 1:
                logger.info(
                    "Start Analysis button disabled - need 2 points for analysis (currently 1)"
                )
            elif num_points == 0 and getattr(self, "current_mode", "") == "tracking":
                logger.info("Start Analysis button disabled - no tracking points available")

        # Check if we're in range selection mode and have 2 points
        if hasattr(self, "range_selection_mode") and self.range_selection_mode:
            current_frame_points = self.viewer_widget.overlay_item.frame_points.get(
                self.viewer_widget.current_frame_index, []
            )
            if len(current_frame_points) >= 2:
                # Two points selected, start tracking workflow
                self.range_selection_mode = False
                if hasattr(self, "tracking_range_limit"):
                    start_frame, end_frame = self.tracking_range_limit
                    self.start_range_tracking_workflow(start_frame, end_frame)

    def toggle_tracking_mode(self):
        """Switch to tracking mode via activity bar"""
        self.activity_bar.set_active_mode("tracking")
        self.on_activity_mode_changed("tracking")

    def create_menus(self):
        """Create menu bar"""
        menubar = self.menuBar()

        # Set menu bar style (VSCode dark theme)
        menubar.setStyleSheet(
            """
            QMenuBar {
                background-color: #2d2d30;
                color: #cccccc;
                font-size: 11px;
                padding: 2px;
                border-bottom: 1px solid #3c3c3c;
            }
            QMenuBar::item {
                padding: 4px 10px;
                background-color: transparent;
            }
            QMenuBar::item:selected {
                background-color: #094771;
                color: white;
            }
            QMenuBar::item:pressed {
                background-color: #007ACC;
                color: white;
            }
            QMenu {
                background-color: #252526;
                color: #cccccc;
                border: 1px solid #3c3c3c;
                font-size: 11px;
            }
            QMenu::item {
                padding: 5px 20px;
            }
            QMenu::item:selected {
                background-color: #094771;
                color: white;
            }
            QMenu::separator {
                height: 1px;
                background: #3c3c3c;
                margin: 5px 0;
            }
        """
        )

        # File menu
        file_menu = menubar.addMenu("&File")

        open_action = QAction("&Open DICOM...", self)
        open_action.setShortcut(QKeySequence.StandardKey.Open)
        open_action.triggered.connect(self.open_dicom)
        file_menu.addAction(open_action)

        open_folder_action = QAction("Open DICOM &Folder...", self)
        open_folder_action.setShortcut("Ctrl+Shift+O")
        open_folder_action.triggered.connect(self.open_dicom_folder)
        file_menu.addAction(open_folder_action)

        # Default DICOM folder action
        open_default_folder_action = QAction("Open &Default DICOM Folder", self)
        open_default_folder_action.setShortcut("Ctrl+D")
        open_default_folder_action.triggered.connect(self.open_default_dicom_folder)
        file_menu.addAction(open_default_folder_action)

        file_menu.addSeparator()

        # Recent files menu
        self.recent_menu = QMenu("&Recent Files", self)
        file_menu.addMenu(self.recent_menu)
        self.update_recent_menu()

        file_menu.addSeparator()

        exit_action = QAction("E&xit", self)
        exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Settings menu
        settings_menu = menubar.addMenu("&Settings")

        # Advanced diameter measurement option
        self.advanced_diameter_action = QAction("Use Advanced Diameter Measurement", self)
        self.advanced_diameter_action.setCheckable(True)
        self.advanced_diameter_action.setChecked(self.qca_analyzer.use_advanced_diameter)
        self.advanced_diameter_action.triggered.connect(self.toggle_advanced_diameter)
        settings_menu.addAction(self.advanced_diameter_action)

        # Help menu
        help_menu = menubar.addMenu("&Help")

        about_action = QAction("&About", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def toggle_advanced_diameter(self):
        """Toggle advanced diameter measurement method"""
        self.qca_analyzer.use_advanced_diameter = self.advanced_diameter_action.isChecked()

        # Show information message
        method = "advanced" if self.qca_analyzer.use_advanced_diameter else "simple"
        QMessageBox.information(
            self,
            "Diameter Measurement Method",
            f"Diameter measurement method changed to: {method.upper()}\n\n"
            f"Advanced method provides:\n"
            f"• Sub-pixel precision\n"
            f"• Multiple measurement techniques\n"
            f"• Adaptive smoothing\n"
            f"• Better accuracy for stenosis detection",
        )

        # Log the change
        logger.info(f"Diameter measurement method changed to: {method}")

    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(
            self,
            "About Coronary Clear Vision",
            "Coronary Clear Vision\n"
            "Version 2.0\n\n"
            "A DICOM viewer for coronary angiography\n"
            "with EKG integration and automated stenosis analysis.",
        )

    def update_zoom_status(self, zoom_level: float):
        """Update zoom level in status bar"""
        self.status_bar.showMessage(f"Zoom: {zoom_level:.0%}", 2000)

    def update_pixel_info(self, info: str):
        """Update pixel info in status bar"""
        self.status_bar.showMessage(info, 2000)

    def update_wl_menu(self):
        """Update window/level presets menu"""

    def load_settings(self):
        """Load application settings"""
        settings = QSettings()

        # Restore window geometry
        geometry = settings.value("window_geometry")
        if geometry:
            self.restoreGeometry(geometry)

        # Check if we should auto-load last DICOM
        last_dicom_path = settings.value("last_dicom_path")
        last_dicom_type = settings.value("last_dicom_type")

        if last_dicom_path and Path(last_dicom_path).exists():
            # Delay auto-load slightly to ensure UI is ready
            QTimer.singleShot(
                500,
                lambda: self.auto_load_last_dicom(
                    last_dicom_path, last_dicom_type, settings.value("last_series_index", 0, int)
                ),
            )

    def save_settings(self):
        """Save application settings"""
        settings = QSettings()
        settings.setValue("window_geometry", self.saveGeometry())

        # Save last opened DICOM file path and type
        if hasattr(self, "last_loaded_path"):
            settings.setValue("last_dicom_path", self.last_loaded_path)
            settings.setValue("last_dicom_type", self.last_loaded_type)  # 'file' or 'folder'

            # If it was a folder with multiple series, save the selected series index
            if hasattr(self, "current_series_index") and hasattr(self, "all_series"):
                settings.setValue("last_series_index", self.current_series_index)

    def auto_load_last_dicom(self, file_path: str, load_type: str, series_index: int = 0):
        """Auto-load the last opened DICOM file/folder on startup"""
        try:
            self.update_status(f"Loading last opened DICOM...")

            if load_type == "file":
                self.load_dicom_file(file_path)
            elif load_type == "folder":
                # Load folder
                folder_loader = DicomFolderLoader()
                if folder_loader.load_folder(file_path):
                    series_list = folder_loader.get_series_list()

                    if series_list:
                        # Store folder loader
                        self.folder_loader = folder_loader
                        self.all_series = series_list

                        # Load the previously selected series if available
                        if 0 <= series_index < len(series_list):
                            self.current_series_index = series_index
                            self._load_series(series_list[series_index])

                            # Enable projection controls if multiple series
                            if len(series_list) > 1:
                                self.prev_projection_button.setEnabled(True)
                                self.next_projection_button.setEnabled(True)
                                self.projection_select_button.setEnabled(True)

            logger.info(f"Auto-loaded last DICOM: {file_path}")

        except Exception as e:
            logger.error(f"Failed to auto-load last DICOM: {e}")
            self.update_status("Failed to load last DICOM")

    def toggle_points_visibility(self, checked):
        """Toggle visibility of points overlay"""
        if hasattr(self.viewer_widget, "overlay_item"):
            self.viewer_widget.overlay_item.show_points = checked
            self.viewer_widget.overlay_item.update()
            self.viewer_widget._request_update()

    def toggle_segmentation_visibility(self, checked):
        """Toggle visibility of segmentation overlay"""
        if hasattr(self.viewer_widget, "overlay_item"):
            self.viewer_widget.overlay_item.show_segmentation = checked
            self.viewer_widget.overlay_item.update()
            self.viewer_widget._request_update()

    def toggle_qca_visibility(self, checked):
        """Toggle visibility of QCA overlay"""
        if hasattr(self.viewer_widget, "overlay_item"):
            self.viewer_widget.overlay_item.show_qca = checked
            self.viewer_widget.overlay_item.update()
            self.viewer_widget._request_update()

    def clear_current_points(self):
        """Clear points from current frame only"""
        if hasattr(self.viewer_widget, "overlay_item"):
            self.viewer_widget.overlay_item.user_points = []
            self.viewer_widget._request_update()
            self.update_status("Current frame points cleared")

    def clear_all_frame_points(self):
        """Clear all tracking points from all frames"""
        reply = QMessageBox.question(
            self,
            "Clear All Frame Points",
            "Are you sure you want to clear all tracking points from all frames?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.viewer_widget.clear_user_points()
            self.update_status("All frame tracking points cleared")

    def clear_all_overlays(self):
        """Clear all overlays including segmentation and QCA for current frame"""
        reply = QMessageBox.question(
            self,
            "Clear Current Overlays",
            "Are you sure you want to clear all overlays from the current frame?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.viewer_widget.clear_overlays()
            self.update_status("Current frame overlays cleared")

    def clear_all_frames_overlays(self):
        """Clear all overlays from all frames"""
        reply = QMessageBox.question(
            self,
            "Clear All Frames Overlays",
            "Are you sure you want to clear all segmentation and QCA overlays from ALL frames?\n\n"
            "This will remove all segmentation masks and QCA results.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            # Clear viewer widget overlays
            self.viewer_widget.clear_overlays()

            # Clear stored results
            if hasattr(self, "batch_segmentation_results"):
                self.batch_segmentation_results.clear()
            if hasattr(self, "batch_qca_results"):
                self.batch_qca_results.clear()
            if hasattr(self.viewer_widget, "frame_segmentation_results"):
                self.viewer_widget.frame_segmentation_results.clear()

            # Clear segmentation widget results if it exists
            if self.segmentation_widget:
                self.segmentation_widget.clear_all_results()

            # Clear QCA widget results if it exists
            if self.qca_widget:
                self.qca_widget.clear_results()

            self.update_status("All frames overlays cleared")

    def resizeEvent(self, event):
        """Handle window resize to update content panel"""
        super().resizeEvent(event)
        # Update content panel size when window is resized
        if hasattr(self, "content_panel"):
            self.update_content_panel_size()

    def closeEvent(self, event):
        """Handle window close"""
        self.save_settings()
        event.accept()

    def track_all_frames(self):
        """Track points through all frames"""
        from PyQt6.QtCore import Qt
        from PyQt6.QtWidgets import QProgressDialog

        # Determine tracking range
        if hasattr(self, "tracking_range_limit") and self.tracking_range_limit:
            # Use pre-defined range (from range selection mode)
            start_frame, end_frame = self.tracking_range_limit
            message = f"Tracking points through frames {start_frame + 1} to {end_frame + 1}..."
            logger.info(f"Using predefined tracking range: {start_frame + 1} to {end_frame + 1}")
        else:
            # Use current beat range instead of all frames
            current_frame = self.viewer_widget.current_frame_index
            start_frame, end_frame = self._get_current_beat_frame_range(current_frame)

            # Start from current frame instead of beat start
            start_frame = current_frame

            message = f"Tracking points through current cardiac cycle (frames {start_frame + 1} to {end_frame + 1})..."
            logger.info(f"Using current beat range: frames {start_frame + 1} to {end_frame + 1}")

        # Create progress dialog
        progress = QProgressDialog(message, "Cancel", 0, 100, self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        def update_progress(current, total):
            progress.setValue(current)
            return not progress.wasCanceled()

        # Perform batch tracking with range limits
        success = self.viewer_widget.track_all_frames(
            progress_callback=update_progress, start_frame=start_frame, end_frame=end_frame
        )

        progress.close()

        if success:
            self.update_status(
                f"Successfully tracked points through frames {start_frame + 1} to {end_frame + 1}"
            )
            # Stop blinking when tracking is completed
            self.stop_track_button_blinking()
            # Refresh current frame display
            self.viewer_widget.display_frame(self.viewer_widget.current_frame_index)
            # Update segmentation widget
            if hasattr(self, "segmentation_widget") and self.segmentation_widget:
                self.segmentation_widget.on_frame_changed(self.viewer_widget.current_frame_index)

            # Directly proceed to sequential processing after tracking
            logger.info(
                f"Tracking complete for frames {start_frame + 1} to {end_frame + 1}, starting sequential processing"
            )
            self.tracked_range = (start_frame, end_frame)
            self.run_sequential_processing(start_frame, end_frame)
        else:
            self.update_status("Failed to track all frames")

    def show_tracking_complete_dialog(self, start_frame: int, end_frame: int):
        """Show dialog after tracking completion to allow editing"""
        from PyQt6.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Tracking Complete")
        msg.setText(f"Tracking completed for frames {start_frame + 1} to {end_frame + 1}.")
        msg.setInformativeText(
            "You can now review and edit the tracked points:\n\n"
            "• Use the frame slider to navigate through frames\n"
            "• Click and drag points to adjust their positions\n"
            "• When satisfied, click 'Proceed to Analysis'"
        )

        # Add custom buttons
        proceed_btn = msg.addButton("Proceed to Analysis", QMessageBox.ButtonRole.ActionRole)
        msg.addButton("Continue Editing", QMessageBox.ButtonRole.RejectRole)

        msg.exec()
        clicked_btn = msg.clickedButton()

        if clicked_btn == proceed_btn:
            # Store the range for later use
            self.tracked_range = (start_frame, end_frame)
            self.show_auto_segmentation_options(start_frame, end_frame)
        else:
            # User wants to edit - store the range and show the proceed button
            self.tracked_range = (start_frame, end_frame)
            self.proceed_to_analysis_btn.setVisible(True)
            self.update_status(
                "You can now edit the tracked points. Use the frame slider to review."
            )

    def on_proceed_to_analysis(self):
        """Handle proceed to analysis button click"""
        if hasattr(self, "tracked_range"):
            start_frame, end_frame = self.tracked_range
            # Hide the button
            self.proceed_to_analysis_btn.setVisible(False)
            # Show auto segmentation options
            self.show_auto_segmentation_options(start_frame, end_frame)
        else:
            self.update_status("No tracked range found. Please track points first.")

    def show_auto_segmentation_options(self, start_frame: int, end_frame: int):
        """Show options for auto segmentation and QCA after tracking"""
        from PyQt6.QtWidgets import QMessageBox

        # Check if we already have segmentation results for this range
        has_existing_segmentation = False
        if hasattr(self, "batch_segmentation_results"):
            # Check if all frames in range have segmentation
            existing_count = sum(
                1
                for frame_idx in range(start_frame, end_frame + 1)
                if frame_idx in self.batch_segmentation_results
                and self.batch_segmentation_results[frame_idx].get("success")
            )
            has_existing_segmentation = existing_count == (end_frame - start_frame + 1)

        msg = QMessageBox(self)
        msg.setWindowTitle("Tracking Complete")

        if has_existing_segmentation:
            msg.setText(
                f"Tracking completed. Found existing segmentation for all frames {start_frame + 1} to {end_frame + 1}."
            )
            msg.setInformativeText("Would you like to run QCA analysis on existing segmentation?")

            # Add custom buttons
            sequential_btn = msg.addButton(
                "Sequential Processing", QMessageBox.ButtonRole.ActionRole
            )
            msg.addButton(QMessageBox.StandardButton.Cancel)

            msg.exec()
            clicked_btn = msg.clickedButton()

            if clicked_btn == sequential_btn:
                self.run_sequential_processing(start_frame, end_frame)
        else:
            msg.setText(f"Ready to analyze frames {start_frame + 1} to {end_frame + 1}.")
            msg.setInformativeText(
                "Sequential Processing will:\n\n"
                "1. Perform vessel segmentation on each frame\n"
                "2. Run QCA analysis to measure stenosis\n"
                "3. Generate comprehensive results\n\n"
                "This process may take several minutes."
            )

            # Add custom buttons
            sequential_btn = msg.addButton(
                "Start Sequential Processing", QMessageBox.ButtonRole.ActionRole
            )
            msg.addButton(QMessageBox.StandardButton.Cancel)

            msg.exec()
            clicked_btn = msg.clickedButton()

            if clicked_btn == sequential_btn:
                self.run_sequential_processing(start_frame, end_frame)

    # Analysis dialogs removed - widgets are now integrated in accordion panel

    def toggle_qca_dialog(self):
        """Switch to QCA analysis tab in accordion"""
        # Switch to QCA mode
        self.activity_bar.set_active_mode("qca")
        self.on_activity_mode_changed("qca")

        # Pass calibration to QCA widget if available
        if self.calibration_factor:
            self.qca_widget.set_calibration(
                self.calibration_factor, getattr(self, "calibration_details", None)
            )

    def toggle_segmentation_dialog(self):
        """Switch to Segmentation tab in accordion"""
        # Find Segmentation tab index
        for i in range(self.content_panel.count()):
            if self.content_panel.itemText(i) == "Segmentation":
                self.content_panel.setCurrentIndex(i)
                break

        # Set frame range in segmentation widget
        total_frames = self.dicom_parser.get_frame_count()
        self.segmentation_widget.set_frame_range(total_frames)

        # Activate segmentation mode
        if hasattr(self.segmentation_widget, "mode_button"):
            if not self.segmentation_widget.mode_button.isChecked():
                self.segmentation_widget.mode_button.setChecked(True)
                self.segmentation_widget.toggle_mode()

    def start_calibration(self):
        """Switch to Calibration tab in accordion"""
        if not self.dicom_parser.dicom_data:
            QMessageBox.warning(self, "No Image", "Please load a DICOM image first")
            return

        # Find Calibration tab index
        for i in range(self.content_panel.count()):
            if self.content_panel.itemText(i) == "Calibration":
                self.content_panel.setCurrentIndex(i)
                break

        # Activate calibration mode
        self.calibration_widget.set_calibration_mode(True)

    def on_calibration_mode_changed(self, enabled: bool):
        """Handle calibration mode changes"""
        self.viewer_widget.set_calibration_mode(enabled)

        if not enabled:
            # Clear any calibration graphics when mode is disabled
            self.viewer_widget.clear_calibration_points()

    def on_viewer_calibration_point_clicked(self, x: int, y: int):
        """Handle calibration point click from viewer"""
        # Pass to calibration widget if in calibration mode
        if self.current_mode == "calibration" and self.calibration_widget.calibration_mode:
            self.calibration_widget.add_point(x, y)

    def on_calibration_completed_widget(self, factor: float, details: dict, dialog=None):
        """Handle calibration completion from widget"""
        # Store calibration values
        self.calibration_factor = factor  # Already in mm/pixel
        self.calibration_details = details

        logger.info(f"=== CALIBRATION RECEIVED IN MAIN WINDOW ===")
        logger.info(f"Calibration factor: {factor:.5f} mm/pixel")
        logger.info(f"Details: {details}")
        logger.info(f"Expected 30px vessel diameter: {30 * factor:.2f}mm")

        # Update viewer with calibration
        if hasattr(self.viewer_widget, "set_calibration"):
            self.viewer_widget.set_calibration(self.calibration_factor)

        # Automatically switch to tracking mode after calibration
        self.activity_bar.set_active_mode("tracking")
        self.on_activity_mode_changed("tracking")

        # Show calibration mask like segmentation
        if "mask" in details and hasattr(self.viewer_widget, "set_segmentation_overlay"):
            # Use segmentation overlay to show calibration area
            overlay_settings = {
                "enabled": True,
                "opacity": 1.0,
                "color": "Green",
                "contour_only": True,
            }
            self.viewer_widget.set_segmentation_overlay(details, overlay_settings)

        # Update status
        status_msg = f"Calibrated: {self.calibration_factor:.5f} mm/pixel"
        self.statusBar().showMessage(status_msg)

        pixels_per_mm = 1.0 / self.calibration_factor if self.calibration_factor > 0 else 0
        QMessageBox.information(
            self,
            "Success",
            f"Calibration applied successfully\n" f"Scale: {pixels_per_mm:.2f} pixels/mm",
        )

        # Clear overlay after 3 seconds
        QTimer.singleShot(3000, lambda: self.viewer_widget.clear_segmentation_graphics())

        # Close dialog after short delay if provided (for backward compatibility)
        if dialog is not None:
            QTimer.singleShot(1000, dialog.close)

    def on_calibration_completed(self, result):
        """Handle calibration completion"""
        # Validate pixels_per_mm before division
        if result.pixels_per_mm <= 0 or result.pixels_per_mm > 1000:  # Sanity check
            logger.error(f"Invalid pixels_per_mm value: {result.pixels_per_mm}")
            QMessageBox.warning(
                self,
                "Calibration Error",
                f"Invalid calibration value detected.\nPlease try calibration again.",
            )
            return

        # Store calibration result
        self.calibration_factor = 1.0 / result.pixels_per_mm  # Convert to mm/pixel
        self.calibration_details = {
            "catheter_width_mm": result.catheter_width_mm,
            "catheter_width_pixels": result.catheter_width_pixels,
            "pixels_per_mm": result.pixels_per_mm,
        }

        # Update viewer with calibration
        if hasattr(self.viewer_widget, "set_calibration"):
            self.viewer_widget.set_calibration(self.calibration_factor)

        # Update status with comparison to fallback
        status_msg = f"Calibrated: {self.calibration_factor:.5f} mm/pixel"
        if self.fallback_calibration is not None:
            diff_percent = abs(
                (self.calibration_factor - self.fallback_calibration)
                / self.fallback_calibration
                * 100
            )
            status_msg += f" (DICOM: {self.fallback_calibration:.5f}, diff: {diff_percent:.1f}%)"
        self.update_status(status_msg)

        # Store in QCA analyzer
        self.qca_analyzer.calibration_factor = self.calibration_factor

    def on_calibration_reset(self):
        """Handle calibration reset"""
        # Clear calibration points but keep in calibration mode
        self.viewer_widget.calibration_points = []
        self.viewer_widget.overlay_item.calibration_line = None
        self.viewer_widget.overlay_item.update()

        # Make sure we stay in calibration mode
        if hasattr(self.viewer_widget, "catheter_size") and hasattr(
            self.viewer_widget, "catheter_diameter_mm"
        ):
            # Stay in calibration mode with same catheter settings
            self.viewer_widget.set_interaction_mode("calibrate")

        # Don't use fallback calibration - it's unreliable
        self.qca_analyzer.calibration_factor = None
        if self.fallback_calibration is not None:
            self.update_status(
                f"Calibration deleted. DICOM pixel spacing ({self.fallback_calibration:.5f} mm/pixel) available but not used - please perform manual calibration"
            )
        else:
            self.update_status("Calibration deleted. No calibration available")

    def extract_and_display_ekg(self):
        """Extract and display EKG data from current DICOM"""
        if self.dicom_parser.dicom_data is None:
            return

        # Try to extract EKG data
        if self.ekg_parser.extract_ekg_from_dicom(self.dicom_parser.dicom_data):
            # Clear cached R-peaks
            if hasattr(self, "_cached_r_peaks"):
                delattr(self, "_cached_r_peaks")

            # Detect R-peaks
            r_peaks = self.ekg_parser.detect_r_peaks()

            # Get quality metrics
            quality_metrics = self.ekg_parser.get_signal_quality()

            # Display EKG data with quality metrics
            self.ekg_viewer.set_ekg_data(
                self.ekg_parser.ekg_data, self.ekg_parser.sampling_rate, r_peaks, quality_metrics
            )

            # Connect R-peak detection to heartbeat counter
            # Disconnect any previous connections first
            try:
                self.ekg_viewer.r_peak_clicked.disconnect()
            except:
                pass
            self.ekg_viewer.r_peak_clicked.connect(self.on_r_peak_detected)

            # Detect cardiac phases for RWS analysis
            cardiac_phases = self.ekg_parser.detect_cardiac_phases()
            if cardiac_phases:
                self.ekg_viewer.set_cardiac_phases(cardiac_phases)

                # Map phases to frames for easier lookup
                if "phases" in cardiac_phases and self.dicom_parser.has_data():
                    frame_rate = self.dicom_parser.get_frame_rate()
                    total_frames = self.dicom_parser.get_frame_count()

                    # Use the phase detector to map phases to frames
                    from src.core.cardiac_phase_detector import CardiacPhaseDetector

                    detector = CardiacPhaseDetector(self.ekg_parser.sampling_rate)
                    frame_phases = detector.map_phases_to_frames(
                        cardiac_phases["phases"], total_frames, frame_rate
                    )
                    cardiac_phases["frame_phases"] = frame_phases
                    logger.info(f"Mapped {len(frame_phases)} phase transitions to frames")

                # Share cardiac phases with viewer widget
                self.viewer_widget.cardiac_phases = cardiac_phases

            # Calculate and display heart rate
            heart_rate = self.ekg_parser.get_heart_rate()
            if heart_rate:
                self.viewer_widget.update_heart_rate(heart_rate)
                status_msg = f"EKG loaded | Heart Rate: {heart_rate:.0f} BPM"
                if quality_metrics:
                    status_msg += f" | Quality: {quality_metrics.get('quality_score', 0):.2f}"
                if cardiac_phases and "statistics" in cardiac_phases:
                    stats = cardiac_phases["statistics"]
                    status_msg += f" | Systole: {stats.get('systole_duration_ms_mean', 0):.0f}ms"
                self.update_status(status_msg)
        else:
            # No EKG data found
            self.ekg_viewer.info_label.setText("No EKG data found in DICOM")

    def extract_frame_timestamps(self):
        """Extract frame timestamps for synchronization"""
        if self.dicom_parser.dicom_data is None:
            return

        ds = self.dicom_parser.dicom_data
        num_frames = self.dicom_parser.num_frames

        # Initialize timestamps array
        self.frame_timestamps = np.zeros(num_frames)

        # Try different methods to get frame times
        if hasattr(ds, "FrameTime") and ds.FrameTime:
            # Frame time in milliseconds
            frame_time_ms = float(ds.FrameTime)
            # Create timestamps for each frame
            self.frame_timestamps = np.arange(num_frames) * (frame_time_ms / 1000.0)

        elif hasattr(ds, "FrameTimeVector") and ds.FrameTimeVector:
            # Direct frame time vector
            try:
                times = [float(t) for t in ds.FrameTimeVector]
                if len(times) == num_frames:
                    self.frame_timestamps = np.array(times) / 1000.0  # Convert to seconds
            except Exception as e:
                logger.warning(f"Ignored exception: {e}")

        elif hasattr(ds, "CineRate") and ds.CineRate:
            # Cine rate (frames per second)
            fps = float(ds.CineRate)
            if fps > 0:
                self.frame_timestamps = np.arange(num_frames) / fps
        else:
            # Default: assume 15 fps (standard for cardiac cine)
            self.frame_timestamps = np.arange(num_frames) / 15.0

    def on_ekg_time_clicked(self, time: float):
        """Handle click on EKG - navigate to corresponding frame"""
        if self.frame_timestamps is None:
            return

        # Find closest frame
        differences = np.abs(self.frame_timestamps - time)
        closest_frame = np.argmin(differences)

        # Navigate to that frame
        self.frame_slider.setValue(closest_frame)

        # Update frame marker on EKG
        self.update_ekg_frame_marker(closest_frame)

    def update_ekg_frame_marker(self, frame_index: int):
        """Update the frame marker on EKG display"""
        if self.frame_timestamps is None or frame_index >= len(self.frame_timestamps):
            return

        # Get timestamp for this frame
        timestamp = self.frame_timestamps[frame_index]

        # Set marker on EKG
        self.ekg_viewer.set_time_marker(timestamp)

    def on_frame_range_selected(self, start_frame: int, end_frame: int):
        """Handle frame range selection from viewer widget"""
        logger.info(f"Frame range selected: {start_frame} to {end_frame}")

        # Store the selected range
        self.selected_frame_range = (start_frame, end_frame)

        # Limit navigation to this range
        self.set_navigation_range(start_frame, end_frame)

        # Navigate to start frame
        self.frame_slider.setValue(start_frame)
        self.viewer_widget.display_frame(start_frame)

        # Check if we have reference points in the start frame
        start_frame_points = self.viewer_widget.overlay_item.frame_points.get(start_frame, [])

        if len(start_frame_points) < 2:
            # No points yet, prompt user to select points
            from PyQt6.QtWidgets import QMessageBox

            QMessageBox.information(
                self,
                "Select Reference Points",
                f"Frame range {start_frame + 1} to {end_frame + 1} selected.\n\n"
                f"Please select 2 reference points on the vessel in this frame,\n"
                f"then use tracking to propagate them through the range.",
            )

            # Enable segmentation mode directly
            if self.segmentation_widget:
                # Enable segmentation mode programmatically
                if not self.segmentation_widget.segmentation_mode:
                    self.segmentation_widget.toggle_mode()

                # Set frame range in segmentation widget
                total_frames = self.dicom_parser.get_frame_count()
                self.segmentation_widget.set_frame_range(total_frames)

            # Set up a flag to track that we're in range selection mode
            self.range_selection_mode = True
            self.tracking_range_limit = (start_frame, end_frame)
        else:
            # Points already exist, ask if user wants to use them or select new ones
            from PyQt6.QtWidgets import QMessageBox

            msg = QMessageBox(self)
            msg.setWindowTitle("Reference Points Found")
            msg.setText(f"Found {len(start_frame_points)} reference points in frame {start_frame}.")
            msg.setInformativeText("Would you like to use these points or select new ones?")

            use_btn = msg.addButton("Use Existing Points", QMessageBox.ButtonRole.AcceptRole)
            new_btn = msg.addButton("Select New Points", QMessageBox.ButtonRole.RejectRole)
            msg.addButton(QMessageBox.StandardButton.Cancel)

            msg.exec()
            clicked_btn = msg.clickedButton()

            if clicked_btn == use_btn:
                # Start tracking workflow
                self.start_range_tracking_workflow(start_frame, end_frame)
            elif clicked_btn == new_btn:
                # Clear existing points and start selection
                self.viewer_widget.clear_all_points_in_frame()

                # Enable segmentation mode directly
                if self.segmentation_widget:
                    if not self.segmentation_widget.segmentation_mode:
                        self.segmentation_widget.toggle_mode()

                    # Set frame range in segmentation widget
                    total_frames = self.dicom_parser.get_frame_count()
                    self.segmentation_widget.set_frame_range(total_frames)

                self.range_selection_mode = True
                self.tracking_range_limit = (start_frame, end_frame)

    def set_navigation_range(self, start_frame: int, end_frame: int):
        """Limit navigation controls to specific frame range"""
        logger.info(f"Setting navigation range to frames {start_frame}-{end_frame}")

        # Store navigation limits
        self.navigation_range = (start_frame, end_frame)

        # Update frame slider range
        self.frame_slider.setMinimum(start_frame)
        self.frame_slider.setMaximum(end_frame)

        # Make sure current value is within range
        current = self.frame_slider.value()
        if current < start_frame:
            self.frame_slider.setValue(start_frame)
        elif current > end_frame:
            self.frame_slider.setValue(end_frame)

        # Update status to show range
        self.update_status(f"Navigation limited to frames {start_frame}-{end_frame}")

    def clear_navigation_range(self):
        """Clear navigation range limits"""
        if hasattr(self, "navigation_range"):
            delattr(self, "navigation_range")

        # Reset frame slider to full range
        if self.dicom_parser.has_data():
            self.frame_slider.setMinimum(0)
            self.frame_slider.setMaximum(self.dicom_parser.num_frames - 1)
            self.update_status("Navigation range cleared")

    def start_range_tracking_workflow(self, start_frame: int, end_frame: int):
        """Start the tracking workflow for the selected range"""
        logger.info(f"Starting tracking workflow for frames {start_frame} to {end_frame}")

        # Store range limits
        self.tracking_range_limit = (start_frame, end_frame)

        # Enable tracking mode
        self.viewer_widget.set_tracking_enabled(True)

        # Log tracking mode activation instead of showing dialog
        logger.info(f"Tracking mode enabled for frames {start_frame + 1} to {end_frame + 1}")

        # Set flag to show completion dialog after tracking
        self.pending_auto_segmentation = True

    def run_sequential_processing(self, start_frame: int, end_frame: int):
        """Run sequential frame-by-frame segmentation and QCA for tracked frames"""
        logger.info(f"Starting sequential processing for frames {start_frame} to {end_frame}")

        # Check calibration - use fallback from DICOM if available
        effective_calibration = self.calibration_factor
        calibration_source = "user calibration"

        if not effective_calibration and self.dicom_parser.pixel_spacing is not None:
            effective_calibration = self.dicom_parser.pixel_spacing
            calibration_source = "DICOM pixel spacing"
            logger.info(
                f"Using DICOM pixel spacing as calibration fallback: {effective_calibration:.5f} mm/pixel"
            )

        if not effective_calibration:
            reply = QMessageBox.question(
                self,
                "No Calibration",
                "No calibration factor available (neither user calibration nor DICOM pixel spacing).\n"
                "Continue with segmentation only?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        else:
            # Log calibration info instead of showing dialog
            logger.info(f"Using {calibration_source}: {effective_calibration:.5f} mm/pixel")

        # Create sequential processor with effective calibration and UI settings
        from src.core.sequential_processor import create_sequential_processor

        # Get UI toggle states
        # Curvature resistant centerline removed - always use default method
        # Light mask limiting removed - always use moderate/tight limiting

        self.sequential_processor = create_sequential_processor(
            self,
            start_frame,
            end_frame,
            effective_calibration,
            use_curvature_resistant_centerline=False,  # Always False
        )

        if self.sequential_processor is None:
            QMessageBox.warning(
                self,
                "No Tracked Frames",
                f"No frames with tracking points found in range {start_frame + 1}-{end_frame + 1}.",
            )
            return

        # Connect signals
        self.sequential_processor.frame_started.connect(self.on_sequential_frame_started)
        self.sequential_processor.segmentation_completed.connect(self.on_sequential_segmentation)
        self.sequential_processor.qca_completed.connect(self.on_sequential_qca)
        self.sequential_processor.frame_completed.connect(self.on_sequential_frame_completed)
        self.sequential_processor.all_completed.connect(self.on_sequential_all_completed)
        self.sequential_processor.error_occurred.connect(self.on_sequential_error)
        self.sequential_processor.progress_updated.connect(self.update_batch_progress)

        # Start processing
        self.sequential_processor.start()

        # Show progress dialog
        tracked_count = len(self.sequential_processor.tracked_frames)
        self.show_batch_progress_dialog("Sequential Processing", tracked_count)

    def show_batch_progress_dialog(self, title: str, total_frames: int):
        """Show progress dialog for batch processing"""
        from PyQt6.QtWidgets import QProgressDialog

        self.batch_progress = QProgressDialog(title, "Cancel", 0, total_frames, self)
        self.batch_progress.setWindowModality(Qt.WindowModality.WindowModal)
        self.batch_progress.setMinimumDuration(0)
        self.batch_progress.setValue(0)
        self.batch_progress.show()

    def update_batch_progress(self, current: int, total: int, message: str):
        """Update batch processing progress"""
        if hasattr(self, "batch_progress"):
            self.batch_progress.setValue(current)
            self.batch_progress.setLabelText(message)

            # Check if cancelled
            if self.batch_progress.wasCanceled():
                if hasattr(self, "sequential_processor"):
                    self.sequential_processor.stop()

    # Sequential processing handlers
    def on_sequential_frame_started(self, frame_idx: int, current: int, total: int):
        """Handle sequential frame start"""
        logger.info(f"Sequential: Starting frame {frame_idx} ({current}/{total})")

        # Update viewer to show current frame
        self.viewer_widget.display_frame(frame_idx)

        # Clear previous overlays when starting a new frame
        if hasattr(self.viewer_widget, "overlay_item"):
            self.viewer_widget.overlay_item.qca_results = None
            self.viewer_widget.overlay_item.segmentation_mask = None
            self.viewer_widget._request_update()

    def on_sequential_segmentation(self, frame_idx: int, result: dict):
        """Handle sequential segmentation completion"""
        # Store segmentation result
        if not hasattr(self.viewer_widget, "frame_segmentation_results"):
            self.viewer_widget.frame_segmentation_results = {}
        self.viewer_widget.frame_segmentation_results[frame_idx] = result

        # Update display if current frame
        if frame_idx == self.viewer_widget.current_frame_index:
            settings = {"enabled": True, "opacity": 1.0, "color": "Red", "contour_only": False}
            self.viewer_widget.set_segmentation_overlay(result, settings)

    def on_sequential_qca(self, frame_idx: int, result: dict):
        """Handle sequential QCA completion"""
        logger.info(f"Sequential: QCA completed for frame {frame_idx}")

        # Add cardiac phase information to the result
        if result.get("success"):
            if hasattr(self.viewer_widget, "cardiac_phases") and self.viewer_widget.cardiac_phases:
                cardiac_phases = self.viewer_widget.cardiac_phases
                logger.info(f"Frame {frame_idx}: Found cardiac_phases data: {bool(cardiac_phases)}")

                # Find phase for this frame
                if "frame_phases" in cardiac_phases:
                    # Use pre-mapped frame phases
                    logger.info(f"Frame {frame_idx}: Using frame_phases mapping")
                    phase_found = False

                    for phase_info in cardiac_phases["frame_phases"]:
                        if phase_info["frame_start"] <= frame_idx <= phase_info["frame_end"]:
                            result["cardiac_phase"] = phase_info["phase"]
                            if "beat_number" in phase_info:
                                result["beat_number"] = phase_info["beat_number"]
                            logger.info(
                                f"Frame {frame_idx}: Added cardiac phase '{result['cardiac_phase']}' "
                                f"({phase_info['phase_name']}) "
                                f"Beat {phase_info.get('beat_number', 'N/A')} "
                                f"(range: {phase_info['frame_start']}-{phase_info['frame_end']})"
                            )
                            phase_found = True
                            break

                    if not phase_found:
                        logger.warning(
                            f"Frame {frame_idx}: No matching cardiac phase found in frame_phases"
                        )
                elif "phases" in cardiac_phases:
                    # Fallback to old method
                    logger.info(f"Frame {frame_idx}: Using old phase detection method")
                    phase_found = False

                    for phase_info in cardiac_phases["phases"]:
                        # Log phase info structure
                        logger.debug(f"Phase info keys: {phase_info.keys()}")

                        if "frame_start" in phase_info and "frame_end" in phase_info:
                            if phase_info["frame_start"] <= frame_idx <= phase_info["frame_end"]:
                                result["cardiac_phase"] = phase_info.get("phase", "")
                                logger.info(
                                    f"Frame {frame_idx}: Added cardiac phase '{result['cardiac_phase']}' "
                                    f"(range: {phase_info['frame_start']}-{phase_info['frame_end']})"
                                )
                                phase_found = True
                                break
                        elif "start_frame" in phase_info and "end_frame" in phase_info:
                            if phase_info["start_frame"] <= frame_idx <= phase_info["end_frame"]:
                                result["cardiac_phase"] = phase_info.get("phase", "")
                                logger.info(
                                    f"Frame {frame_idx}: Added cardiac phase '{result['cardiac_phase']}' "
                                    f"(range: {phase_info['start_frame']}-{phase_info['end_frame']})"
                                )
                                phase_found = True
                                break

                    if not phase_found:
                        logger.warning(f"Frame {frame_idx}: No matching cardiac phase found")
                else:
                    logger.warning(f"Frame {frame_idx}: No 'phases' key in cardiac_phases data")
            else:
                logger.warning(
                    f"Frame {frame_idx}: No cardiac_phases data available in viewer_widget"
                )

                # Try to get cardiac phase from frame metadata if available
                if hasattr(self.dicom_parser, "get_frame_cardiac_phase"):
                    frame_phase = self.dicom_parser.get_frame_cardiac_phase(frame_idx)
                    if frame_phase:
                        result["cardiac_phase"] = frame_phase
                        logger.info(
                            f"Frame {frame_idx}: Got cardiac phase from DICOM metadata: '{frame_phase}'"
                        )
                else:
                    # If no cardiac phase data available, try to estimate based on frame position
                    # This is a simple estimation for demonstration
                    total_frames = self.dicom_parser.get_frame_count()
                    if total_frames > 0:
                        # Simple cyclic pattern estimation
                        cycle_position = (
                            frame_idx % 30
                        ) / 30.0  # Assume ~30 frames per cardiac cycle
                        if cycle_position < 0.3:
                            result["cardiac_phase"] = "d2"  # End-diastole
                        elif cycle_position < 0.5:
                            result["cardiac_phase"] = "s1"  # Early-systole
                        elif cycle_position < 0.7:
                            result["cardiac_phase"] = "s2"  # End-systole
                        else:
                            result["cardiac_phase"] = "d1"  # Mid-diastole
                        logger.info(
                            f"Frame {frame_idx}: Estimated cardiac phase: '{result['cardiac_phase']}'"
                        )

        # Display frame if we're on the current frame being processed
        current_frame = self.viewer_widget.current_frame_index
        if frame_idx == current_frame:
            # Update QCA overlay with centerline and points
            if result.get("success") and hasattr(self.viewer_widget, "set_qca_overlay"):
                # Get overlay settings from QCA widget
                settings = (
                    self.qca_widget.get_overlay_settings()
                    if hasattr(self, "qca_widget")
                    else {
                        "show_measurements": True,
                        "show_stenosis": True,
                        "show_reference": True,
                        "show_centerline": True,
                        "color": "Yellow",
                    }
                )

                # Add enabled flag
                settings["enabled"] = True

                # Set the QCA overlay with centerline and stenosis points
                self.viewer_widget.set_qca_overlay(result, settings)

    def on_sequential_frame_completed(self, frame_idx: int, seg_result: dict, qca_result: dict):
        """Handle sequential frame completion"""
        logger.info(f"Sequential: Frame {frame_idx} fully processed")

        # Store frame-specific results for later viewing
        if not hasattr(self.viewer_widget, "frame_qca_results"):
            self.viewer_widget.frame_qca_results = {}
        if not hasattr(self.viewer_widget, "frame_segmentation_results"):
            self.viewer_widget.frame_segmentation_results = {}

        self.viewer_widget.frame_qca_results[frame_idx] = qca_result
        self.viewer_widget.frame_segmentation_results[frame_idx] = seg_result

    def on_sequential_all_completed(self, seg_results: dict, qca_results: dict):
        """Handle sequential processing completion"""
        logger.info("🎯 [SIGNAL DEBUG] on_sequential_all_completed called!")
        logger.info(f"🎯 [SIGNAL DEBUG] seg_results: {len(seg_results)} frames")
        logger.info(f"🎯 [SIGNAL DEBUG] qca_results: {len(qca_results)} frames")
        print("🎯 [SIGNAL DEBUG] on_sequential_all_completed called!")
        print(f"🎯 [SIGNAL DEBUG] seg_results: {len(seg_results)} frames")
        print(f"🎯 [SIGNAL DEBUG] qca_results: {len(qca_results)} frames")

        # Store results
        self.batch_segmentation_results = seg_results
        self.batch_qca_results = qca_results
        self.sequential_qca_results = qca_results  # Store for RWS analysis

        # Calculate global reference and recalculate all QCA results with it
        logger.info("=== POST-PROCESSING: Applying global reference to current beat ===")
        self._recalculate_beat_with_global_reference(seg_results, qca_results)

        # Close progress dialog
        if hasattr(self, "batch_progress"):
            self.batch_progress.close()

        # Update QCA widget with all results
        if qca_results and hasattr(self, "qca_widget"):
            # Switch to QCA panel
            self.activity_bar.set_active_mode("qca")
            self.on_activity_mode_changed("qca")

            # Update sequential summary in QCA widget
            self.qca_widget.update_sequential_summary(qca_results)

        # Show completion message
        if qca_results:
            QMessageBox.information(
                self,
                "Sequential Processing Complete",
                f"Processed {len(qca_results)} frames successfully.\n"
                f"Results are shown in the QCA panel.",
            )
        else:
            QMessageBox.information(
                self,
                "Sequential Processing Complete",
                f"Processed {len(seg_results)} frames successfully.",
            )

    def _recalculate_beat_with_global_reference(self, seg_results: dict, qca_results: dict):
        """Calculate global reference diameter and recalculate all QCA results with it"""
        try:
            logger.info("🔄 [POST-PROCESS DEBUG] _recalculate_beat_with_global_reference CALLED!")
            print("🔄 [POST-PROCESS DEBUG] _recalculate_beat_with_global_reference CALLED!")

            # NOTE: This function recalculates stenosis percentages using global reference
            # but preserves frame-specific diameter measurements
            logger.info(
                "📊 Starting global reference calculation to standardize stenosis measurements"
            )

            logger.info(
                f"[POST-PROCESS DEBUG] Starting recalculation with QCA analyzer: {self.qca_analyzer}"
            )

            # Debug: Check inputs
            logger.info(
                f"[POST-PROCESS DEBUG] seg_results type: {type(seg_results)}, length: {len(seg_results) if seg_results else 0}"
            )
            logger.info(
                f"[POST-PROCESS DEBUG] qca_results type: {type(qca_results)}, length: {len(qca_results) if qca_results else 0}"
            )

            if not qca_results:
                logger.warning("No QCA results available for global reference calculation")
                return

            # Collect frame-based reference diameters from QCA results
            frame_reference_diameters = []
            logger.info(
                f"[POST-PROCESS DEBUG] Checking {len(qca_results)} QCA results for frame reference diameters..."
            )

            for frame_idx, result in qca_results.items():
                logger.info(
                    f"[POST-PROCESS DEBUG] Frame {frame_idx}: success={result.get('success')}, has_frame_reference={'frame_reference_diameter' in result if result else False}"
                )

                if result.get("success") and "frame_reference_diameter" in result:
                    frame_ref = result["frame_reference_diameter"]
                    if frame_ref is not None and frame_ref > 0:
                        frame_reference_diameters.append(frame_ref)
                        logger.info(
                            f"[POST-PROCESS DEBUG] Frame {frame_idx}: Frame reference diameter: {frame_ref:.2f}mm"
                        )
                    else:
                        logger.info(
                            f"[POST-PROCESS DEBUG] Frame {frame_idx}: Invalid frame reference diameter"
                        )
                else:
                    logger.info(
                        f"[POST-PROCESS DEBUG] Frame {frame_idx}: No frame_reference_diameter found in result"
                    )

            if len(frame_reference_diameters) == 0:
                logger.warning("No valid frame reference diameters found in QCA results")
                return

            # Calculate global reference diameter (75th percentile of frame references)
            import numpy as np

            global_reference_diameter = np.percentile(frame_reference_diameters, 75)
            logger.info(
                f"[POST-PROCESS DEBUG] Calculated global_reference_diameter: {global_reference_diameter}"
            )

            logger.info(f"=== GLOBAL REFERENCE DIAMETER CALCULATED ===")
            logger.info(f"Total frame reference diameters: {len(frame_reference_diameters)}")
            logger.info(
                f"Frame reference range: {np.min(frame_reference_diameters):.2f} - {np.max(frame_reference_diameters):.2f} mm"
            )
            logger.info(
                f"75th percentile of frame references (Global Reference): {global_reference_diameter:.2f} mm"
            )

            # Now recalculate all frames with global reference
            logger.info("Recalculating all QCA results with global reference...")
            updated_qca_results = {}

            for frame_idx, seg_result in seg_results.items():
                try:
                    logger.info(
                        f"[POST-PROCESS DEBUG] Processing frame {frame_idx}: seg_result_success={seg_result.get('success') if seg_result else False}"
                    )
                    if seg_result and seg_result.get("success"):
                        # Get tracking points for this frame
                        tracking_points = None
                        # Try to get tracking points from segmentation result first
                        if "reference_points" in seg_result:
                            tracking_points = seg_result["reference_points"]
                            logger.info(
                                f"[POST-PROCESS DEBUG] Frame {frame_idx}: Got tracking points from seg_result"
                            )
                        else:
                            # Fallback to overlay_item
                            if hasattr(self, "viewer_widget"):
                                overlay_item = getattr(self.viewer_widget, "overlay_item", None)
                                if overlay_item and hasattr(overlay_item, "get_frame_points"):
                                    tracking_points = overlay_item.get_frame_points(frame_idx)
                                elif overlay_item and hasattr(overlay_item, "frame_points"):
                                    tracking_points = overlay_item.frame_points.get(frame_idx, [])
                            logger.info(
                                f"[POST-PROCESS DEBUG] Frame {frame_idx}: Got tracking points from overlay_item"
                            )

                        logger.info(
                            f"[POST-PROCESS DEBUG] Frame {frame_idx}: tracking_points={len(tracking_points) if tracking_points else 0}"
                        )
                        if tracking_points and len(tracking_points) >= 2:
                            # Get original frame
                            logger.info(
                                f"[POST-PROCESS DEBUG] Frame {frame_idx}: Getting frame data..."
                            )
                            frame = self.dicom_parser.get_frame(
                                frame_idx
                            )  # Use dicom_parser instead
                            if frame is None:
                                logger.error(
                                    f"[POST-PROCESS DEBUG] Frame {frame_idx}: Failed to get frame data"
                                )
                                continue

                            # Use existing segmentation result which already has centerline from tracking points
                            # Recalculate QCA with global reference (direct analyze_vessel call)
                            logger.info(
                                f"[POST-PROCESS DEBUG] About to call analyze_vessel with global_reference_diameter: {global_reference_diameter}"
                            )

                            # Get centerline from original QCA result (it was calculated with tracking points)
                            original_qca = qca_results[frame_idx]
                            if "centerline" not in original_qca:
                                logger.error(
                                    f"[POST-PROCESS DEBUG] Frame {frame_idx}: No centerline in original QCA result"
                                )
                                continue

                            np.array(original_qca["centerline"])

                            # Ensure QCA analyzer has correct calibration factor
                            if hasattr(self, "calibration_factor") and self.calibration_factor:
                                self.qca_analyzer.calibration_factor = self.calibration_factor
                                logger.info(
                                    f"[POST-PROCESS DEBUG] Set calibration_factor to: {self.calibration_factor}"
                                )
                            else:
                                logger.warning(
                                    f"[POST-PROCESS DEBUG] No calibration_factor available"
                                )

                            # IMPORTANT: Use analyze_from_angiopy to preserve frame-specific diameters
                            # analyze_vessel with same centerline produces same results
                            logger.info(
                                f"[POST-PROCESS DEBUG] Re-analyzing frame {frame_idx} with global reference"
                            )

                            # Get original frame for diameter measurement
                            original_image = frame

                            # Extract proximal and distal points from tracking points
                            proximal_point = (
                                tracking_points[0] if len(tracking_points) > 0 else None
                            )
                            distal_point = tracking_points[-1] if len(tracking_points) > 1 else None

                            # Use segmentation-based method to preserve frame-specific diameters
                            # Global reference is only used for stenosis calculation
                            use_gradient_method = False
                            gradient_method = "second_derivative"

                            updated_result = self.qca_analyzer.analyze_from_angiopy(
                                seg_result,
                                original_image=original_image,
                                proximal_point=proximal_point,
                                distal_point=distal_point,
                                tracked_points=(
                                    tracking_points if len(tracking_points) >= 2 else None
                                ),
                                use_tracked_centerline=True,
                                use_gradient_method=use_gradient_method,
                                gradient_method=gradient_method,
                                global_reference_diameter=global_reference_diameter,
                            )

                            if updated_result.get("success"):
                                # Add frame metadata from original QCA result
                                original_qca = qca_results[frame_idx]
                                updated_result["frame_index"] = frame_idx
                                updated_result["display_frame_id"] = (
                                    frame_idx + 1
                                )  # UI shows 1-based

                                # Copy cardiac phase info from original result
                                if "cardiac_phase" in original_qca:
                                    updated_result["cardiac_phase"] = original_qca["cardiac_phase"]
                                if "frame_type" in original_qca:
                                    updated_result["frame_type"] = original_qca["frame_type"]
                                if "phase" in original_qca:
                                    updated_result["phase"] = original_qca["phase"]

                                # Get cardiac phase from main window's phase mapping if available
                                if hasattr(self, "phase_frames") and frame_idx in self.phase_frames:
                                    phase_info = self.phase_frames[frame_idx]
                                    updated_result["cardiac_phase"] = phase_info.get("phase", "")
                                    updated_result["frame_type"] = phase_info.get("type", "")

                                updated_qca_results[frame_idx] = updated_result
                                logger.info(
                                    f"Frame {frame_idx}: Updated with global reference - "
                                    f"MLD: {updated_result.get('mld', 0):.3f}mm, "
                                    f"Ref: {global_reference_diameter:.2f}mm, "
                                    f"Stenosis: {updated_result['percent_stenosis']:.1f}%"
                                )

                except Exception as e:
                    logger.error(f"Failed to update QCA for frame {frame_idx}: {e}")
                    continue

            if updated_qca_results:
                # Replace old results with updated ones
                self.batch_qca_results.update(updated_qca_results)
                self.sequential_qca_results.update(updated_qca_results)

                # Update QCA widget with recalculated results
                if hasattr(self, "qca_widget"):
                    self.qca_widget.update_sequential_summary(updated_qca_results)
                    logger.info(
                        f"Updated QCA widget with {len(updated_qca_results)} recalculated results"
                    )

                logger.info(f"=== BEAT RECALCULATION COMPLETE ===")
                logger.info(
                    f"Successfully updated {len(updated_qca_results)} QCA results with global reference"
                )
                logger.info(
                    f"All frames in current beat now use consistent global reference: {global_reference_diameter:.2f}mm"
                )

        except Exception as e:
            logger.error(f"Global reference calculation failed: {e}")

    def get_cardiac_phase_for_frame(self, frame_idx: int) -> dict:
        """Get cardiac phase information for a specific frame"""
        try:
            # Check if we have cardiac phase analysis
            if (
                hasattr(self, "ekg_viewer")
                and self.ekg_viewer
                and hasattr(self.ekg_viewer, "phase_analysis")
            ):
                phase_analysis = self.ekg_viewer.phase_analysis
                if phase_analysis and hasattr(phase_analysis, "phases"):
                    for phase_name, frame_list in phase_analysis.phases.items():
                        if frame_idx in frame_list:
                            return {"phase": phase_name, "type": phase_name}

            # Fallback: check if frame is in specific beat ranges
            if hasattr(self, "phase_transitions") and self.phase_transitions:
                for transition in self.phase_transitions:
                    if (
                        transition.get("start_frame", 0)
                        <= frame_idx
                        <= transition.get("end_frame", 0)
                    ):
                        return {
                            "phase": transition.get("phase", "unknown"),
                            "type": transition.get("phase", "unknown"),
                        }

            return {}
        except Exception as e:
            logger.warning(f"Failed to get cardiac phase for frame {frame_idx}: {e}")
            return {}

    def on_sequential_error(self, frame_idx: int, error_msg: str):
        """Handle sequential processing error"""
        if frame_idx >= 0:
            logger.error(f"Sequential: Error on frame {frame_idx}: {error_msg}")
        else:
            logger.error(f"Sequential: General error: {error_msg}")

    def show_qca_results_summary_OLD(self, qca_results: dict):
        """DEPRECATED: Now integrated into QCA widget"""
        # This method is no longer used - results are shown in QCA widget

    def _show_qca_results_summary_dialog(self, qca_results: dict):
        """Show a summary of QCA results in a table dialog (backup method)"""
        from PyQt6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QTableWidget,
            QTableWidgetItem,
            QPushButton,
            QHBoxLayout,
            QHeaderView,
        )

        # Create dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("QCA Analysis Results")
        dialog.setModal(False)  # Non-modal so user can interact with main window
        dialog.resize(1200, 600)

        layout = QVBoxLayout()

        # Create table widget
        table = QTableWidget()
        table.setRowCount(len(qca_results))

        # Define columns
        columns = [
            "Frame",
            "Cardiac Phase",
            "Stenosis %",
            "MLD (mm)",
            "RVD (mm)",
            "Lesion Length (mm)",
            "Area Stenosis %",
            "MLA (mm²)",
            "RVA (mm²)",
        ]
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)

        # Cardiac phase mapping (phase transition names)
        cardiac_phases = {
            "d2": "End-diastole",
            "s1": "Early-systole",
            "s2": "End-systole",
            "d1": "Mid-diastole",
        }

        # Get cardiac phase data if available
        cardiac_phase_data = {}
        if (
            hasattr(self, "ekg_parser")
            and self.ekg_parser
            and hasattr(self.ekg_parser, "cardiac_phases")
        ):
            # Calculate frame times and map to cardiac phases
            frame_rate = (
                self.dicom_parser.get_frame_rate() if hasattr(self, "dicom_parser") else 30.0
            )
            frame_duration = 1.0 / frame_rate  # seconds per frame

            # Debug logging
            logger.info(
                f"Frame rate: {frame_rate} fps, Frame duration: {frame_duration:.4f} seconds"
            )

            for frame_idx in qca_results.keys():
                frame_time = frame_idx * frame_duration
                phase = self.ekg_parser.get_phase_at_time(frame_time)
                if phase:
                    cardiac_phase_data[frame_idx] = phase
                    logger.debug(f"Frame {frame_idx}: time={frame_time:.3f}s, phase={phase}")
                else:
                    logger.warning(
                        f"No phase found for frame {frame_idx} at time {frame_time:.3f}s"
                    )

        # Populate table
        row = 0
        for frame_idx in sorted(qca_results.keys()):
            result = qca_results[frame_idx]
            if not result.get("success"):
                continue

            # Frame number
            table.setItem(row, 0, QTableWidgetItem(str(frame_idx)))

            # Cardiac phase
            phase = cardiac_phase_data.get(frame_idx, "")
            phase_text = cardiac_phases.get(phase, phase)
            table.setItem(row, 1, QTableWidgetItem(phase_text))

            # QCA parameters - map the actual keys from QCA results
            params = [
                ("percent_stenosis", 2, "{:.1f}"),
                ("mld", 3, "{:.2f}"),  # MLD in mm
                ("reference_diameter", 4, "{:.2f}"),  # RVD in mm
                ("lesion_length", 5, "{:.2f}"),  # Lesion length in mm
                ("percent_area_stenosis", 6, "{:.1f}"),  # Area stenosis %
                ("mla", 7, "{:.2f}"),  # MLA in mm²
                ("reference_area", 8, "{:.2f}"),  # RVA in mm²
            ]

            for param, col, fmt in params:
                value = result.get(param, 0)
                if value is not None:
                    table.setItem(row, col, QTableWidgetItem(fmt.format(value)))
                else:
                    table.setItem(row, col, QTableWidgetItem("N/A"))

            row += 1

        # Adjust table properties
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        # Connect row selection to frame navigation
        def on_row_selected():
            selected_row = table.currentRow()
            if selected_row >= 0:
                frame_item = table.item(selected_row, 0)
                if frame_item:
                    frame_idx = int(frame_item.text())
                    self.navigate_to_frame(frame_idx)

        table.itemSelectionChanged.connect(on_row_selected)

        layout.addWidget(table)

        # Add summary statistics
        stenosis_values = [
            r["percent_stenosis"]
            for r in qca_results.values()
            if r.get("success") and "percent_stenosis" in r
        ]

        if stenosis_values:
            from PyQt6.QtWidgets import QLabel

            summary_text = (
                f"Summary: Average stenosis: {sum(stenosis_values)/len(stenosis_values):.1f}%, "
                f"Maximum: {max(stenosis_values):.1f}%, "
                f"Minimum: {min(stenosis_values):.1f}%"
            )
            summary_label = QLabel(summary_text)
            layout.addWidget(summary_label)

        # Add buttons
        button_layout = QHBoxLayout()

        export_btn = QPushButton("Export CSV")
        export_btn.clicked.connect(
            lambda: self.export_qca_results_csv(qca_results, cardiac_phase_data, cardiac_phases)
        )
        button_layout.addWidget(export_btn)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        dialog.setLayout(layout)

        # Store dialog reference to prevent garbage collection
        self.qca_results_dialog = dialog
        dialog.show()

    def export_qca_results_csv(
        self, qca_results: dict, cardiac_phase_data: dict, cardiac_phases: dict
    ):
        """Export QCA results to CSV file"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import csv

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export QCA Results", "", "CSV Files (*.csv)"
        )

        if file_path:
            with open(file_path, "w", newline="") as csvfile:
                fieldnames = [
                    "Frame",
                    "Cardiac Phase",
                    "Stenosis %",
                    "MLD (mm)",
                    "RVD (mm)",
                    "Lesion Length (mm)",
                    "Area Stenosis %",
                    "MLA (mm²)",
                    "RVA (mm²)",
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for frame_idx in sorted(qca_results.keys()):
                    result = qca_results[frame_idx]
                    if result.get("success"):
                        phase = cardiac_phase_data.get(frame_idx, "")
                        phase_text = cardiac_phases.get(phase, phase)

                        writer.writerow(
                            {
                                "Frame": frame_idx,
                                "Cardiac Phase": phase_text,
                                "Stenosis %": f"{result.get('percent_stenosis', 0):.1f}",
                                "MLD (mm)": f"{result.get('mld', 0):.3f}",
                                "RVD (mm)": f"{result.get('reference_diameter', 0):.2f}",
                                "Lesion Length (mm)": f"{result.get('lesion_length', 0):.2f}",
                                "Area Stenosis %": f"{result.get('percent_area_stenosis', 0):.1f}",
                                "MLA (mm²)": f"{result.get('mla', 0):.2f}",
                                "RVA (mm²)": f"{result.get('reference_area', 0):.2f}",
                            }
                        )

            QMessageBox.information(self, "Export Complete", f"Results exported to {file_path}")

    def navigate_to_frame(self, frame_index: int):
        """Navigate to a specific frame"""
        self.frame_slider.setValue(frame_index)
        self.viewer_widget.display_frame(frame_index)

    def start_sequential_processing(self):
        """Start sequential processing directly on tracked frames"""
        if not self.dicom_parser.has_data():
            QMessageBox.warning(self, "No Data", "Please load a DICOM file first.")
            return

        # Check if we have tracking points
        tracked_frames = []
        if hasattr(self.viewer_widget, "overlay_item") and hasattr(
            self.viewer_widget.overlay_item, "frame_points"
        ):
            tracked_frames = sorted(self.viewer_widget.overlay_item.frame_points.keys())

        if not tracked_frames:
            QMessageBox.warning(
                self, "No Tracking Points", "Please track vessel points first using Tracking Mode."
            )
            return

        # Get the range of tracked frames
        start_frame = min(tracked_frames)
        end_frame = max(tracked_frames)

        # Show confirmation dialog
        msg = QMessageBox(self)
        msg.setWindowTitle("Sequential Processing")
        msg.setText(f"Found {len(tracked_frames)} tracked frames")
        msg.setInformativeText(
            f"Process frames {start_frame + 1} to {end_frame + 1}?\n\n"
            f"This will run segmentation and QCA analysis on all tracked frames."
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg.setIcon(QMessageBox.Icon.Question)

        if msg.exec() == QMessageBox.StandardButton.Yes:
            self.run_sequential_processing(start_frame, end_frame)

    def start_qca_analysis(self):
        """Start QCA analysis"""
        if not self.dicom_parser.has_data():
            QMessageBox.warning(self, "No Data", "Please load a DICOM file first.")
            return

        # Switch to QCA mode
        self.activity_bar.set_active_mode("qca")
        self.on_activity_mode_changed("qca")

        # Pass calibration to QCA widget
        if self.calibration_factor:
            self.qca_widget.set_calibration(
                self.calibration_factor, getattr(self, "calibration_details", None)
            )

    def start_qca_from_segmentation(self, segmentation_result):
        """Start QCA analysis using AngioPy segmentation data"""
        # Switch to QCA mode
        self.activity_bar.set_active_mode("qca")
        self.on_activity_mode_changed("qca")

        # Pass calibration to QCA widget
        if self.calibration_factor:
            self.qca_widget.set_calibration(
                self.calibration_factor, getattr(self, "calibration_details", None)
            )

        # Check if reference points are provided in segmentation result
        reference_points = segmentation_result.get("reference_points", None)
        if reference_points:
            logger.info(f"Starting QCA with reference points: {reference_points}")
            # Store reference points in QCA widget for later use
            self.qca_widget.reference_points = reference_points
        else:
            logger.info("Starting QCA without reference point limitation")
            self.qca_widget.reference_points = None

        # Start QCA analysis with segmentation data
        self.qca_widget.start_analysis(segmentation_result)

    def start_segmentation(self):
        """Start vessel segmentation"""
        if not self.dicom_parser.has_data():
            QMessageBox.warning(self, "No Data", "Please load a DICOM file first.")
            return

        # Show tracking tips when starting segmentation
        QMessageBox.information(
            self,
            "Tracking Tips",
            "For optimal tracking results:\n\n"
            "• Select vessel bifurcations (Y-shaped junctions)\n"
            "• Choose stent edges or markers\n"
            "• Pick catheter tips or distinctive features\n"
            "• Avoid uniform vessel segments\n\n"
            "CoTracker3 will track the exact points you select.",
        )

        # Switch to Segmentation mode
        self.activity_bar.set_active_mode("segmentation")
        self.on_activity_mode_changed("segmentation")

        # Set frame range in segmentation widget
        total_frames = self.dicom_parser.get_frame_count()
        self.segmentation_widget.set_frame_range(total_frames)

        # Auto-enable segmentation mode if not already enabled
        if not self.segmentation_widget.segmentation_mode:
            self.segmentation_widget.segmentation_mode_btn.setChecked(True)
            self.segmentation_widget.toggle_segmentation_mode()

    def save_current_image(self):
        """Save current frame as image"""
        if not self.dicom_parser.has_data():
            QMessageBox.warning(self, "No Data", "No image to save.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Image",
            str(Path.home() / "coronary_image.png"),
            "PNG Files (*.png);;JPEG Files (*.jpg *.jpeg);;All Files (*)",
        )

        if file_path:
            # Get current pixmap from viewer
            if hasattr(self.viewer_widget, "get_current_pixmap"):
                pixmap = self.viewer_widget.get_current_pixmap()
                if pixmap:
                    pixmap.save(file_path)
                    self.update_status(f"Image saved to {file_path}")

    def export_video(self):
        """Export frames as video"""
        if not self.dicom_parser.has_data():
            QMessageBox.warning(self, "No Data", "No frames to export.")
            return

        QMessageBox.information(
            self,
            "Export Video",
            "Video export functionality will be implemented in a future version.",
        )

    def export_analysis_report(self, format_type="pdf"):
        """Export analysis report in various formats"""
        if format_type == "xlsx":
            # Use comprehensive XLSX export
            self.export_comprehensive_xlsx_report()
            return
        elif format_type == "pdf":
            self.export_pdf_report()
            return
        elif format_type == "txt":
            self.export_txt_report()
            return

        # Fallback: show preview dialog
        report_text = self._generate_analysis_report()
        self._show_report_preview(report_text)

    def update_analysis_report(self):
        """Update analysis report - now just triggers UI updates"""
        # Since we removed the report panel, just update any status displays

    def _generate_analysis_report(self) -> str:
        """Generate comprehensive analysis report text"""
        from datetime import datetime

        report_lines = []

        # Header
        report_lines.append("=" * 80)
        report_lines.append("CORONARY ANALYSIS COMPREHENSIVE REPORT")
        report_lines.append("=" * 80)
        report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append("")

        # Patient Information
        patient_name = ""
        if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.dicom_data:
            dicom_data = self.dicom_parser.dicom_data
            if hasattr(dicom_data, "PatientName"):
                patient_name = str(getattr(dicom_data, "PatientName", ""))

        if patient_name:
            report_lines.append("PATIENT INFORMATION")
            report_lines.append("-" * 40)
            report_lines.append(f"Patient Name: {patient_name}")
            if hasattr(dicom_data, "PatientID"):
                report_lines.append(f"Patient ID: {getattr(dicom_data, 'PatientID', 'N/A')}")
            if hasattr(dicom_data, "StudyDescription"):
                report_lines.append(f"Study: {getattr(dicom_data, 'StudyDescription', 'N/A')}")
            report_lines.append("")

        # Calibration Information
        report_lines.append("CALIBRATION INFORMATION")
        report_lines.append("-" * 40)
        if hasattr(self, "calibration_factor") and self.calibration_factor:
            report_lines.append(f"Calibration Factor: {self.calibration_factor:.5f} mm/pixel")
            if hasattr(self, "calibration_details"):
                report_lines.append(
                    f"Method: {self.calibration_details.get('method', 'Manual catheter calibration')}"
                )
                report_lines.append(
                    f"Catheter Size: {self.calibration_details.get('catheter_size', 'Unknown')}"
                )
        else:
            report_lines.append("Calibration: Not performed")

        if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.pixel_spacing:
            report_lines.append(
                f"DICOM Pixel Spacing: {self.dicom_parser.pixel_spacing:.5f} mm/pixel"
            )
        report_lines.append("")

        # QCA Analysis Results
        report_lines.append("QCA ANALYSIS RESULTS")
        report_lines.append("-" * 40)

        qca_results = None
        if hasattr(self, "qca_widget") and hasattr(self.qca_widget, "analysis_result"):
            qca_results = self.qca_widget.analysis_result
        elif hasattr(self, "sequential_qca_results"):
            qca_results = self.sequential_qca_results

        if qca_results and isinstance(qca_results, dict):
            valid_results = []
            for frame_idx, result in qca_results.items():
                if (
                    isinstance(frame_idx, int)
                    and isinstance(result, dict)
                    and result.get("success")
                ):
                    valid_results.append((frame_idx, result))

            if valid_results:
                report_lines.append(f"Total Frames Analyzed: {len(valid_results)}")
                report_lines.append("")
                report_lines.append("Frame  | Stenosis(%) | MLD(mm) | Ref.Dia(mm) | Lesion Len(mm)")
                report_lines.append("-" * 65)

                for frame_idx, result in sorted(valid_results):
                    stenosis = result.get("percent_stenosis", 0)
                    mld = result.get("mld", 0) or result.get("minimum_diameter", 0)
                    ref_dia = result.get("reference_diameter", 0)
                    lesion_len = result.get("lesion_length", 0)

                    report_lines.append(
                        f"{frame_idx:5d}  | {stenosis:8.1f}    | {mld:6.2f}  | {ref_dia:9.2f}   | {lesion_len:10.2f}"
                    )
            else:
                report_lines.append("No valid QCA results available")
        else:
            report_lines.append("No QCA analysis performed")
        report_lines.append("")

        # RWS Analysis Results
        report_lines.append("RWS ANALYSIS RESULTS")
        report_lines.append("-" * 40)

        # Get RWS results using the same logic as Excel export
        rws_results = None
        if hasattr(self, "rws_enhanced_results") and self.rws_enhanced_results:
            rws_results = self.rws_enhanced_results
        elif hasattr(self, "rws_enhanced_widget") and hasattr(
            self.rws_enhanced_widget, "analysis_results"
        ):
            rws_results = self.rws_enhanced_widget.analysis_results
        elif hasattr(self, "rws_widget") and hasattr(self.rws_widget, "results"):
            rws_results = self.rws_widget.results

        if rws_results and isinstance(rws_results, dict):
            # Check if this is Enhanced RWS results format
            if "stenosis_rws" in rws_results:
                stenosis_rws = rws_results["stenosis_rws"]
                rws_val = stenosis_rws.get("rws_stenosis", 0)

                report_lines.append(f"Stenosis RWS: {rws_val:.2f}%")
                report_lines.append("Analysis Method: Enhanced RWS Analysis with outlier detection")

                if "mld_values" in stenosis_rws:
                    mld_values = stenosis_rws["mld_values"]
                    if mld_values:
                        max_mld = max(mld_values.values())
                        min_mld = min(mld_values.values())
                        max_phase = stenosis_rws.get("max_mld_phase", "unknown")
                        min_phase = stenosis_rws.get("min_mld_phase", "unknown")

                        report_lines.append(f"Maximum MLD: {max_mld:.2f}mm (Phase: {max_phase})")
                        report_lines.append(f"Minimum MLD: {min_mld:.2f}mm (Phase: {min_phase})")
                        report_lines.append(f"MLD Variation: {max_mld - min_mld:.2f}mm")
            else:
                # Standard RWS format
                rws_val = rws_results.get("rws_percentage", 0)
                report_lines.append(f"RWS: {rws_val:.2f}%")
                if "max_diameter" in rws_results:
                    report_lines.append(f"Maximum MLD: {rws_results['max_diameter']:.2f}mm")
                if "min_diameter" in rws_results:
                    report_lines.append(f"Minimum MLD: {rws_results['min_diameter']:.2f}mm")
                if "method" in rws_results:
                    report_lines.append(f"Method: {rws_results['method']}")

            # Clinical interpretation
            rws_val = rws_results.get("stenosis_rws", {}).get("rws_stenosis", 0) or rws_results.get(
                "rws_percentage", 0
            )
            if rws_val > 12:
                report_lines.append(
                    "Clinical Interpretation: HIGH RWS (>12%) - Potential plaque vulnerability"
                )
            else:
                report_lines.append(
                    "Clinical Interpretation: NORMAL RWS (<12%) - Stable plaque characteristics"
                )
        else:
            report_lines.append("No RWS analysis performed")

        report_lines.append("")
        report_lines.append("RWS Formula: (MLDmax - MLDmin) / MLDmin × 100%")
        report_lines.append("")

        # Footer
        report_lines.append("-" * 80)
        report_lines.append("Report generated by Coronary Clear Vision")
        report_lines.append("-" * 80)

        return "\n".join(report_lines)

    def _show_report_preview(self, report_text):
        """Show report preview dialog"""
        from PyQt6.QtWidgets import QTextEdit, QVBoxLayout

        dialog = QDialog(self)
        dialog.setWindowTitle("Analysis Report Preview")
        dialog.resize(800, 600)

        layout = QVBoxLayout(dialog)
        text_edit = QTextEdit()
        text_edit.setPlainText(report_text)
        text_edit.setReadOnly(True)
        text_edit.setFont(QFont("Courier", 10))  # Monospace font for better formatting
        layout.addWidget(text_edit)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)

        dialog.exec()

    def export_pdf_report(self):
        """Export analysis report as PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from PyQt6.QtWidgets import QFileDialog
            from datetime import datetime

            # Get patient name for filename
            patient_name = ""
            if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.dicom_data:
                dicom_data = self.dicom_parser.dicom_data
                if hasattr(dicom_data, "PatientName"):
                    patient_name_raw = str(getattr(dicom_data, "PatientName", ""))
                    if patient_name_raw:
                        patient_name = self._clean_filename(patient_name_raw)

            # Create filename with patient name
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if patient_name:
                suggested_filename = f"coronary_analysis_{patient_name}_{timestamp}.pdf"
            else:
                suggested_filename = f"coronary_analysis_report_{timestamp}.pdf"

            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save PDF Report", suggested_filename, "PDF Files (*.pdf)"
            )

            if not file_path:
                return

            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center alignment
            )
            story.append(Paragraph("CORONARY ANALYSIS COMPREHENSIVE REPORT", title_style))
            story.append(Spacer(1, 12))

            # Generation info
            story.append(
                Paragraph(
                    f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 12))

            # Patient Information
            if (
                patient_name
                and hasattr(self, "dicom_parser")
                and self.dicom_parser
                and self.dicom_parser.dicom_data
            ):
                dicom_data = self.dicom_parser.dicom_data
                story.append(Paragraph("<b>PATIENT INFORMATION</b>", styles["Heading2"]))

                patient_data = []
                if hasattr(dicom_data, "PatientName"):
                    patient_data.append(
                        ["Patient Name:", str(getattr(dicom_data, "PatientName", "N/A"))]
                    )
                if hasattr(dicom_data, "PatientID"):
                    patient_data.append(
                        ["Patient ID:", str(getattr(dicom_data, "PatientID", "N/A"))]
                    )
                if hasattr(dicom_data, "StudyDescription"):
                    patient_data.append(
                        ["Study:", str(getattr(dicom_data, "StudyDescription", "N/A"))]
                    )

                if patient_data:
                    patient_table = Table(patient_data)
                    patient_table.setStyle(
                        TableStyle(
                            [
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ]
                        )
                    )
                    story.append(patient_table)
                    story.append(Spacer(1, 12))

            # Calibration Information
            story.append(Paragraph("<b>CALIBRATION INFORMATION</b>", styles["Heading2"]))
            cal_data = []

            if hasattr(self, "calibration_factor") and self.calibration_factor:
                cal_data.append(["Calibration Factor:", f"{self.calibration_factor:.5f} mm/pixel"])
                if hasattr(self, "calibration_details"):
                    cal_data.append(
                        [
                            "Method:",
                            self.calibration_details.get("method", "Manual catheter calibration"),
                        ]
                    )
                    cal_data.append(
                        ["Catheter Size:", self.calibration_details.get("catheter_size", "Unknown")]
                    )
            else:
                cal_data.append(["Calibration:", "Not performed"])

            if (
                hasattr(self, "dicom_parser")
                and self.dicom_parser
                and self.dicom_parser.pixel_spacing
            ):
                cal_data.append(
                    ["DICOM Pixel Spacing:", f"{self.dicom_parser.pixel_spacing:.5f} mm/pixel"]
                )

            cal_table = Table(cal_data)
            cal_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(cal_table)
            story.append(Spacer(1, 12))

            # QCA Results
            story.append(Paragraph("<b>QCA ANALYSIS RESULTS</b>", styles["Heading2"]))

            qca_results = None
            if hasattr(self, "sequential_qca_results"):
                qca_results = self.sequential_qca_results

            if qca_results and isinstance(qca_results, dict):
                valid_results = []
                for frame_idx, result in qca_results.items():
                    if (
                        isinstance(frame_idx, int)
                        and isinstance(result, dict)
                        and result.get("success")
                    ):
                        valid_results.append((frame_idx, result))

                if valid_results:
                    story.append(
                        Paragraph(
                            f"<b>Total Frames Analyzed:</b> {len(valid_results)}", styles["Normal"]
                        )
                    )
                    story.append(Spacer(1, 6))

                    # QCA Table
                    qca_table_data = [
                        [
                            "Frame",
                            "Stenosis (%)",
                            "MLD (mm)",
                            "Ref. Diameter (mm)",
                            "Lesion Length (mm)",
                        ]
                    ]
                    for frame_idx, result in sorted(valid_results):
                        stenosis = result.get("percent_stenosis", 0)
                        mld = result.get("mld", 0) or result.get("minimum_diameter", 0)
                        ref_dia = result.get("reference_diameter", 0)
                        lesion_len = result.get("lesion_length", 0)
                        qca_table_data.append(
                            [
                                str(frame_idx),
                                f"{stenosis:.1f}",
                                f"{mld:.2f}",
                                f"{ref_dia:.2f}",
                                f"{lesion_len:.2f}",
                            ]
                        )

                    qca_table = Table(qca_table_data)
                    qca_table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 10),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(qca_table)
                else:
                    story.append(Paragraph("No valid QCA results available", styles["Normal"]))
            else:
                story.append(Paragraph("No QCA analysis performed", styles["Normal"]))

            story.append(Spacer(1, 12))

            # RWS Results
            story.append(Paragraph("<b>RWS ANALYSIS RESULTS</b>", styles["Heading2"]))

            # Get RWS results
            rws_results = None
            if hasattr(self, "rws_enhanced_results") and self.rws_enhanced_results:
                rws_results = self.rws_enhanced_results
            elif hasattr(self, "rws_enhanced_widget") and hasattr(
                self.rws_enhanced_widget, "analysis_results"
            ):
                rws_results = self.rws_enhanced_widget.analysis_results
            elif hasattr(self, "rws_widget") and hasattr(self.rws_widget, "results"):
                rws_results = self.rws_widget.results

            if rws_results and isinstance(rws_results, dict):
                rws_data = []

                if "stenosis_rws" in rws_results:
                    # Enhanced RWS format
                    stenosis_rws = rws_results["stenosis_rws"]
                    rws_val = stenosis_rws.get("rws_stenosis", 0)

                    rws_data.append(["Stenosis RWS:", f"{rws_val:.2f}%"])
                    rws_data.append(
                        ["Analysis Method:", "Enhanced RWS Analysis with outlier detection"]
                    )

                    if "mld_values" in stenosis_rws and stenosis_rws["mld_values"]:
                        mld_values = stenosis_rws["mld_values"]
                        max_mld = max(mld_values.values())
                        min_mld = min(mld_values.values())
                        max_phase = stenosis_rws.get("max_mld_phase", "unknown")
                        min_phase = stenosis_rws.get("min_mld_phase", "unknown")

                        rws_data.append(["Maximum MLD:", f"{max_mld:.2f}mm (Phase: {max_phase})"])
                        rws_data.append(["Minimum MLD:", f"{min_mld:.2f}mm (Phase: {min_phase})"])
                        rws_data.append(["MLD Variation:", f"{max_mld - min_mld:.2f}mm"])
                else:
                    # Standard RWS format
                    rws_val = rws_results.get("rws_percentage", 0)
                    rws_data.append(["RWS:", f"{rws_val:.2f}%"])
                    if "method" in rws_results:
                        rws_data.append(["Method:", rws_results["method"]])

                # Clinical interpretation
                rws_val = rws_results.get("stenosis_rws", {}).get(
                    "rws_stenosis", 0
                ) or rws_results.get("rws_percentage", 0)
                if rws_val > 12:
                    interpretation = "HIGH RWS (>12%) - Potential plaque vulnerability"
                    interp_color = colors.red
                else:
                    interpretation = "NORMAL RWS (<12%) - Stable plaque characteristics"
                    interp_color = colors.green

                rws_data.append(["Clinical Interpretation:", interpretation])

                rws_table = Table(rws_data)
                rws_table.setStyle(
                    TableStyle(
                        [
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            (
                                "TEXTCOLOR",
                                (1, -1),
                                (1, -1),
                                interp_color,
                            ),  # Color the interpretation
                        ]
                    )
                )
                story.append(rws_table)
            else:
                story.append(Paragraph("No RWS analysis performed", styles["Normal"]))

            story.append(Spacer(1, 12))
            story.append(
                Paragraph("<b>RWS Formula:</b> (MLDmax - MLDmin) / MLDmin × 100%", styles["Normal"])
            )
            story.append(Spacer(1, 20))

            # Footer
            story.append(Paragraph("Report generated by Coronary Clear Vision", styles["Normal"]))

            # Build PDF
            doc.build(story)

            QMessageBox.information(
                self, "Export Successful", f"PDF report exported successfully:\n{file_path}"
            )

        except Exception as e:
            logger.error(f"Failed to export PDF report: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF report:\n{str(e)}")

    def export_txt_report(self):
        """Export analysis report as text file"""
        try:
            from PyQt6.QtWidgets import QFileDialog
            from datetime import datetime

            # Get patient name for filename
            patient_name = ""
            if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.dicom_data:
                dicom_data = self.dicom_parser.dicom_data
                if hasattr(dicom_data, "PatientName"):
                    patient_name_raw = str(getattr(dicom_data, "PatientName", ""))
                    if patient_name_raw:
                        patient_name = self._clean_filename(patient_name_raw)

            # Create filename with patient name
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if patient_name:
                suggested_filename = f"coronary_analysis_{patient_name}_{timestamp}.txt"
            else:
                suggested_filename = f"coronary_analysis_report_{timestamp}.txt"

            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Text Report", suggested_filename, "Text Files (*.txt)"
            )

            if not file_path:
                return

            # Generate report text
            report_text = self._generate_analysis_report()

            # Save to file
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(report_text)

            QMessageBox.information(
                self, "Export Successful", f"Text report exported successfully:\n{file_path}"
            )

        except Exception as e:
            logger.error(f"Failed to export text report: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export text report:\n{str(e)}")

    def toggle_segmentation_overlay(self):
        """Toggle segmentation overlay visibility"""
        if hasattr(self.viewer_widget, "toggle_segmentation_overlay"):
            # Toggle the current state
            current_state = getattr(self.viewer_widget, "segmentation_overlay_enabled", False)
            self.viewer_widget.toggle_segmentation_overlay(not current_state)

    def toggle_qca_overlay(self):
        """Toggle QCA overlay visibility"""
        if hasattr(self.viewer_widget, "toggle_qca_overlay"):
            # Toggle the current state
            current_state = getattr(self.viewer_widget, "qca_overlay_enabled", False)
            self.viewer_widget.toggle_qca_overlay(not current_state)

    def update_qca_overlay(self, results):
        """Update QCA overlay with analysis results"""
        # Get overlay settings from QCA widget first
        settings = {}
        if hasattr(self, "qca_widget"):
            settings = self.qca_widget.get_overlay_settings()

        # Set QCA results with proper settings
        if hasattr(self.viewer_widget, "set_qca_overlay"):
            settings["enabled"] = True
            self.viewer_widget.set_qca_overlay(results, settings)
        # Update report with QCA results
        self.update_analysis_report()

    def on_segmentation_mode_changed(self, enabled: bool):
        """Handle segmentation mode changes"""
        self.viewer_widget.set_segmentation_mode(enabled)

        if not enabled:
            # Clear any segmentation graphics when mode is disabled
            self.viewer_widget.clear_segmentation_graphics()

    def on_segmentation_overlay_changed(self, settings: dict):
        """Handle segmentation overlay settings change"""
        self.update_segmentation_overlay_settings(settings)

    def update_segmentation_overlay(self, result: dict):
        """Update segmentation overlay with results"""
        if hasattr(self.viewer_widget, "set_segmentation_overlay"):
            settings = self.segmentation_widget.get_overlay_settings()
            self.viewer_widget.set_segmentation_overlay(result, settings)

    def on_viewer_segmentation_point_clicked(self, x: int, y: int):
        """Handle segmentation point click from viewer"""
        # Only forward to segmentation widget if it's in segmentation mode
        if hasattr(self, "segmentation_widget") and self.segmentation_widget.segmentation_mode:
            self.segmentation_widget.add_point(x, y)

    def update_qca_overlay_settings(self, settings):
        """Update QCA overlay settings"""
        if hasattr(self.viewer_widget, "update_qca_overlay_settings"):
            self.viewer_widget.update_qca_overlay_settings(settings)

    def update_segmentation_overlay_settings(self, settings):
        """Update segmentation overlay settings"""
        if hasattr(self.viewer_widget, "update_segmentation_overlay_settings"):
            self.viewer_widget.update_segmentation_overlay_settings(settings)

    def on_calibration_overlay_changed(self, overlay_settings: dict):
        """Handle calibration overlay visibility changes"""
        if not hasattr(self, "viewer_widget") or not self.viewer_widget:
            logger.warning("No viewer widget available for calibration overlay")
            return

        logger.info(f"🎨 Calibration overlay settings changed: {overlay_settings}")
        logger.info(f"Method: {overlay_settings.get('method', 'unknown')}")
        logger.info(f"Result data keys: {list(overlay_settings.get('result_data', {}).keys())}")

        # Clear all calibration overlays if requested
        if overlay_settings.get("clear_all", False):
            if hasattr(self.viewer_widget, "clear_calibration_overlays"):
                self.viewer_widget.clear_calibration_overlays()
            return

        # Update calibration overlays based on settings
        method = overlay_settings.get("method", "angiopy")
        result_data = overlay_settings.get("result_data", {})

        if method == "angiopy":
            # Clear automatic overlays for AngioPy method
            if hasattr(self.viewer_widget, "clear_calibration_centerline_overlay"):
                self.viewer_widget.clear_calibration_centerline_overlay()
            if hasattr(self.viewer_widget, "clear_calibration_diameter_overlay"):
                self.viewer_widget.clear_calibration_diameter_overlay()

            # Handle AngioPy mask overlay
            show_mask = overlay_settings.get("show_angiopy_mask", False)
            if show_mask and "mask" in result_data:
                if hasattr(self.viewer_widget, "set_calibration_mask_overlay"):
                    self.viewer_widget.set_calibration_mask_overlay(result_data["mask"])
            else:
                if hasattr(self.viewer_widget, "clear_calibration_mask_overlay"):
                    self.viewer_widget.clear_calibration_mask_overlay()

        elif method == "automatic":
            # Clear AngioPy mask overlay for automatic method
            if hasattr(self.viewer_widget, "clear_calibration_mask_overlay"):
                self.viewer_widget.clear_calibration_mask_overlay()

            # Handle automatic method overlays
            show_centerline = overlay_settings.get("show_auto_centerline", False)
            show_diameters = overlay_settings.get("show_auto_diameters", False)

            if show_centerline and "centerline" in result_data:
                if hasattr(self.viewer_widget, "set_calibration_centerline_overlay"):
                    self.viewer_widget.set_calibration_centerline_overlay(result_data["centerline"])
            else:
                if hasattr(self.viewer_widget, "clear_calibration_centerline_overlay"):
                    self.viewer_widget.clear_calibration_centerline_overlay()

            if show_diameters:
                logger.info(f"🔍 Diameter overlay requested - checking data availability:")
                logger.info(f"  - show_diameters: {show_diameters}")
                logger.info(f"  - 'centerline' in result_data: {'centerline' in result_data}")
                logger.info(
                    f"  - 'diameter_measurements' in result_data: {'diameter_measurements' in result_data}"
                )

                if "centerline" in result_data and "diameter_measurements" in result_data:
                    # Use actual diameter measurements along centerline
                    centerline = result_data["centerline"]
                    diameters = result_data["diameter_measurements"]

                    logger.info(f"  - Centerline length: {len(centerline)}")
                    logger.info(f"  - Diameters length: {len(diameters)}")

                    if len(centerline) == len(diameters):
                        if hasattr(self.viewer_widget, "set_calibration_diameter_overlay"):
                            logger.info("✅ Setting calibration diameter overlay")
                            self.viewer_widget.set_calibration_diameter_overlay(
                                centerline, diameters
                            )
                        else:
                            logger.error(
                                "❌ Viewer widget has no set_calibration_diameter_overlay method"
                            )
                    else:
                        logger.warning(
                            f"❌ Centerline length ({len(centerline)}) != diameters length ({len(diameters)})"
                        )
                else:
                    logger.warning("❌ Missing centerline or diameter_measurements in result_data")
                    if hasattr(self.viewer_widget, "clear_calibration_diameter_overlay"):
                        self.viewer_widget.clear_calibration_diameter_overlay()
            else:
                logger.info("🔍 Diameter overlay not requested - clearing")
                if hasattr(self.viewer_widget, "clear_calibration_diameter_overlay"):
                    self.viewer_widget.clear_calibration_diameter_overlay()

    def show_shortcuts(self):
        """Show keyboard shortcuts"""
        shortcuts = """
        <h3>Keyboard Shortcuts</h3>
        <table>
        <tr><td><b>Ctrl+O</b></td><td>Open DICOM file</td></tr>
        <tr><td><b>←/→</b></td><td>Previous/Next frame</td></tr>
        <tr><td><b>Home/End</b></td><td>First/Last frame</td></tr>
        <tr><td><b>Space</b></td><td>Play/Pause</td></tr>
        <tr><td><b>Ctrl++/-</b></td><td>Zoom In/Out</td></tr>
        <tr><td><b>F</b></td><td>Fit to window</td></tr>
        <tr><td><b>Mouse Wheel</b></td><td>Zoom</td></tr>
        <tr><td><b>Mouse Drag</b></td><td>Pan</td></tr>
        <tr><td><b>Right Mouse Drag</b></td><td>Adjust Window/Level</td></tr>
        </table>
        """
        QMessageBox.information(self, "Keyboard Shortcuts", shortcuts)

    def keyPressEvent(self, event):
        """Handle keyboard shortcuts - works in all modes"""
        # Always handle playback controls regardless of current mode
        handled = False

        if event.key() == Qt.Key.Key_Left:
            current = self.frame_slider.value()
            if current > 0:
                self.frame_slider.setValue(current - 1)
            handled = True

        elif event.key() == Qt.Key.Key_Right:
            current = self.frame_slider.value()
            if current < self.frame_slider.maximum():
                self.frame_slider.setValue(current + 1)
            handled = True

        elif event.key() == Qt.Key.Key_Home:
            self.frame_slider.setValue(0)
            # Reset heartbeat counter when going to start
            self.viewer_widget.reset_heartbeat_overlay()
            handled = True

        elif event.key() == Qt.Key.Key_End:
            self.frame_slider.setValue(self.frame_slider.maximum())
            handled = True

        elif event.key() == Qt.Key.Key_Space:
            if self.play_button.isEnabled():
                self.toggle_play()
            handled = True

        # If we didn't handle it, pass to parent
        if not handled:
            super().keyPressEvent(event)
        else:
            # Accept the event to prevent propagation
            event.accept()

    def add_to_recent(self, file_path: str):
        """Add file to recent files list"""
        settings = QSettings()
        recent = settings.value("recent_files", [])
        if not isinstance(recent, list):
            recent = []

        # Remove if already exists
        if file_path in recent:
            recent.remove(file_path)

        # Add to front
        recent.insert(0, file_path)

        # Keep only last 10
        recent = recent[:10]

        settings.setValue("recent_files", recent)
        self.update_recent_menu()

    def update_recent_menu(self):
        """Update recent files menu"""
        self.recent_menu.clear()

        settings = QSettings()
        recent = settings.value("recent_files", [])
        if not isinstance(recent, list):
            recent = []

        for file_path in recent:
            if os.path.exists(file_path):
                action = QAction(os.path.basename(file_path), self)
                action.triggered.connect(lambda checked, fp=file_path: self.load_dicom_file(fp))
                self.recent_menu.addAction(action)

        if self.recent_menu.isEmpty():
            action = QAction("(No recent files)", self)
            action.setEnabled(False)
            self.recent_menu.addAction(action)

    def reset_layout(self):
        """Reset the entire application layout"""
        # Clear saved state
        settings = QSettings()
        settings.remove("window_state")

        # No dock positions to reset

        # Reset window size
        self.resize(1280, 720)  # Compact size for 1920x1080 screen

        # Center window on screen
        from PyQt6.QtWidgets import QApplication

        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)

        # Show message
        QMessageBox.information(self, "Layout Reset", "Layout has been reset to default.")

    def check_ekg_sync_status(self):
        """Check and update ECG-DICOM synchronization status"""
        if not hasattr(self, "dicom_parser") or not self.dicom_parser.has_data():
            return

        if not hasattr(self, "ekg_parser") or self.ekg_parser.ekg_data is None:
            return

        # Calculate video duration
        num_frames = self.dicom_parser.get_frame_count()
        frame_rate = self.dicom_parser.get_frame_rate()
        if frame_rate > 0:
            video_duration = num_frames / frame_rate
        else:
            return

        # Calculate ECG duration
        num_samples = len(self.ekg_parser.ekg_data)
        sampling_rate = self.ekg_parser.sampling_rate
        if sampling_rate > 0:
            ecg_duration = num_samples / sampling_rate
        else:
            return

        # Check sync status
        diff_ms = abs(video_duration - ecg_duration) * 1000
        is_synced = diff_ms < 1.0  # 1ms tolerance

        # Update heartbeat overlay
        self.viewer_widget.update_heartbeat_sync_status(is_synced)

    def on_r_peak_detected(self, peak_index: int):
        """Handle R-peak detection for heartbeat counting"""
        # Trigger heartbeat animation
        self.viewer_widget.trigger_heartbeat()

        # If playing, sync with current frame
        if hasattr(self, "frame_timestamps") and hasattr(self, "ekg_parser"):
            # Convert peak index to time
            peak_time = peak_index / self.ekg_parser.sampling_rate

            # Find corresponding frame
            current_frame = self.frame_slider.value()
            if current_frame < len(self.frame_timestamps):
                frame_time = self.frame_timestamps[current_frame]

                # Log synchronization for debugging
                time_diff = abs(peak_time - frame_time)
                if time_diff < 0.1:  # Within 100ms
                    logger.debug(
                        f"R-peak at {peak_time:.3f}s matches frame {current_frame} at {frame_time:.3f}s"
                    )

    def check_r_peak_at_frame(self, frame_index: int):
        """Check if current frame corresponds to an R-peak and update beat counter"""
        if not hasattr(self, "ekg_parser") or self.ekg_parser.ekg_data is None:
            return

        if not hasattr(self, "frame_timestamps") or frame_index >= len(self.frame_timestamps):
            return

        # Get R-peaks if not already stored
        if not hasattr(self, "_cached_r_peaks"):
            self._cached_r_peaks = self.ekg_parser.detect_r_peaks()

        if self._cached_r_peaks is None or len(self._cached_r_peaks) == 0:
            return

        # Get current frame time
        frame_time = self.frame_timestamps[frame_index]

        # Calculate which beat this frame belongs to
        current_beat = self._calculate_current_beat(frame_time)
        total_beats = len(self._cached_r_peaks)

        # Update heartbeat overlay with current beat
        if current_beat > 0:
            self.viewer_widget.set_current_beat(current_beat, total_beats)

        # Check each R-peak for animation trigger
        for peak_idx in self._cached_r_peaks:
            peak_time = peak_idx / self.ekg_parser.sampling_rate

            # If frame time is close to R-peak time (within 20ms)
            if abs(peak_time - frame_time) < 0.02:
                # Trigger heartbeat animation
                self.viewer_widget.trigger_heartbeat()
                break

    def _calculate_current_beat(self, frame_time: float) -> int:
        """Calculate which beat number the current frame time belongs to"""
        if not hasattr(self, "_cached_r_peaks") or self._cached_r_peaks is None:
            return 0

        # Convert R-peak indices to times
        peak_times = self._cached_r_peaks / self.ekg_parser.sampling_rate

        # If before first R-peak, we're in beat 1
        if frame_time < peak_times[0]:
            return 1

        # Find which beat interval we're in
        # Each R-peak starts a new beat
        for i in range(len(peak_times) - 1):
            if peak_times[i] <= frame_time < peak_times[i + 1]:
                return i + 1

        # If after last R-peak, we're in the last beat
        if frame_time >= peak_times[-1]:
            return len(peak_times)

        return 0

    def _get_current_beat_frame_range(self, current_frame: int) -> tuple:
        """Get the frame range for the current cardiac beat"""
        try:
            # If no ECG data, return reasonable default (1 second worth of frames)
            if (
                not hasattr(self, "ekg_parser")
                or not self.ekg_parser
                or not hasattr(self, "_cached_r_peaks")
            ):
                frame_rate = self.dicom_parser.get_frame_rate() or 30.0
                frames_per_beat = int(frame_rate)  # Assume ~1 second per beat
                end_frame = min(current_frame + frames_per_beat, self.dicom_parser.num_frames - 1)
                logger.info(
                    f"No ECG data - using default beat range: frames {current_frame} to {end_frame}"
                )
                return (current_frame, end_frame)

            # Get frame time
            frame_rate = self.dicom_parser.get_frame_rate() or 30.0
            frame_time = current_frame / frame_rate

            # Get R-peak times
            peak_times = self._cached_r_peaks / self.ekg_parser.sampling_rate

            # Find current beat boundaries
            beat_start_time = 0.0
            beat_end_time = float("inf")

            # Find which beat interval we're in
            if frame_time < peak_times[0]:
                # Before first R-peak
                beat_start_time = 0.0
                beat_end_time = peak_times[0] if len(peak_times) > 0 else frame_time + 1.0
            else:
                # Find the beat interval
                for i in range(len(peak_times) - 1):
                    if peak_times[i] <= frame_time < peak_times[i + 1]:
                        beat_start_time = peak_times[i]
                        beat_end_time = peak_times[i + 1]
                        break
                else:
                    # After last R-peak
                    if frame_time >= peak_times[-1]:
                        beat_start_time = peak_times[-1]
                        # Estimate beat end (add average beat duration)
                        if len(peak_times) > 1:
                            avg_beat_duration = np.mean(np.diff(peak_times))
                            beat_end_time = beat_start_time + avg_beat_duration
                        else:
                            beat_end_time = beat_start_time + 1.0  # Default 1 second

            # Convert times to frames
            start_frame = max(0, int(beat_start_time * frame_rate))
            end_frame = min(int(beat_end_time * frame_rate), self.dicom_parser.num_frames - 1)

            # Ensure we have at least current frame to end of beat
            if start_frame > current_frame:
                start_frame = current_frame

            logger.info(
                f"Current beat range: frames {start_frame} to {end_frame} "
                f"(times: {beat_start_time:.2f}s to {beat_end_time:.2f}s)"
            )

            return (start_frame, end_frame)

        except Exception as e:
            logger.error(f"Error calculating beat range: {e}")
            # Fallback to reasonable default
            frame_rate = self.dicom_parser.get_frame_rate() or 30.0
            frames_per_beat = int(frame_rate)  # ~1 second
            end_frame = min(current_frame + frames_per_beat, self.dicom_parser.num_frames - 1)
            return (current_frame, end_frame)

    def start_track_button_blinking(self):
        """Start blinking the Start Analysis button"""
        if not self.track_button_timer.isActive():
            self.track_button_timer.start(600)  # Blink every 600ms

    def stop_track_button_blinking(self):
        """Stop blinking the Start Analysis button"""
        self.track_button_timer.stop()
        # Reset button style to normal
        self.track_all_button.setStyleSheet(
            """
            QPushButton { 
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
                background-color: #2E7D32;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """
        )
        self.track_button_blink_state = False

    def toggle_track_button_style(self):
        """Toggle the Start Analysis button style for blinking effect"""
        self.track_button_blink_state = not self.track_button_blink_state

        if self.track_button_blink_state:
            # Bright green highlight style - eye-catching
            self.track_all_button.setStyleSheet(
                """
                QPushButton { 
                    font-size: 14px;
                    font-weight: bold;
                    padding: 10px;
                    background-color: #4CAF50;
                    color: white;
                    border-radius: 5px;
                    border: 2px solid #81C784;
                }
                QPushButton:hover {
                    background-color: #66BB6A;
                }
            """
            )
        else:
            # Darker green normal state
            self.track_all_button.setStyleSheet(
                """
                QPushButton { 
                    font-size: 14px;
                    font-weight: bold;
                    padding: 10px;
                    background-color: #2E7D32;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #388E3C;
                }
            """
            )

    def export_comprehensive_xlsx_report(self):
        """Export comprehensive analysis report as XLSX with multiple sheets"""
        from openpyxl import Workbook
        from datetime import datetime

        try:
            # Get file path from user
            from PyQt6.QtWidgets import QFileDialog

            # Get patient name from DICOM metadata if available
            patient_name = ""
            if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.dicom_data:
                dicom_data = self.dicom_parser.dicom_data
                if hasattr(dicom_data, "PatientName"):
                    patient_name_raw = str(getattr(dicom_data, "PatientName", ""))
                    if patient_name_raw:
                        # Clean patient name for filename (remove invalid characters)
                        patient_name = self._clean_filename(patient_name_raw)

            # Create filename with patient name if available
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if patient_name:
                suggested_filename = f"coronary_analysis_{patient_name}_{timestamp}.xlsx"
            else:
                suggested_filename = f"coronary_analysis_report_{timestamp}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Comprehensive Analysis Report",
                suggested_filename,
                "Excel Files (*.xlsx)",
            )

            if not file_path:
                return

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Create and populate sheets
            self._create_calibration_sheet(wb)
            self._create_qca_sheet(wb)
            self._create_rws_sheet(wb)
            self._create_dicom_metadata_sheet(wb)

            # Save workbook
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Comprehensive analysis report exported successfully:\n{file_path}",
            )

        except Exception as e:
            logger.error(f"Failed to export XLSX report: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export XLSX report:\n{str(e)}")

    def _create_calibration_sheet(self, wb):
        """Create calibration data sheet"""
        from openpyxl.styles import Font

        ws = wb.create_sheet("Calibration")

        # Header
        header_font = Font(bold=True, size=14)
        ws["A1"] = "Calibration Information"
        ws["A1"].font = header_font

        row = 3

        # Calibration factor
        ws[f"A{row}"] = "Calibration Factor (mm/pixel):"
        if hasattr(self, "calibration_factor") and self.calibration_factor:
            ws[f"B{row}"] = f"{self.calibration_factor:.5f}"
        else:
            ws[f"B{row}"] = "Not calibrated"
        row += 1

        # Calibration method
        ws[f"A{row}"] = "Calibration Method:"
        if hasattr(self, "calibration_details"):
            ws[f"B{row}"] = self.calibration_details.get("method", "Unknown")
        else:
            ws[f"B{row}"] = "Manual catheter calibration"
        row += 1

        # Catheter size
        ws[f"A{row}"] = "Catheter Size:"
        if hasattr(self, "calibration_details"):
            ws[f"B{row}"] = self.calibration_details.get("catheter_size", "Unknown")
        else:
            ws[f"B{row}"] = "Unknown"
        row += 1

        # DICOM fallback calibration
        if hasattr(self, "fallback_calibration") and self.fallback_calibration:
            ws[f"A{row}"] = "DICOM Calibration (mm/pixel):"
            ws[f"B{row}"] = f"{self.fallback_calibration:.5f}"
            row += 1

        # Pixel spacing from DICOM
        if hasattr(self, "dicom_parser") and self.dicom_parser and self.dicom_parser.pixel_spacing:
            ws[f"A{row}"] = "DICOM Pixel Spacing (mm/pixel):"
            ws[f"B{row}"] = f"{self.dicom_parser.pixel_spacing:.5f}"
            row += 1

        # Auto-fit columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_qca_sheet(self, wb):
        """Create QCA analysis data sheet"""
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("QCA Analysis")

        # Header
        header_font = Font(bold=True, size=14)
        ws["A1"] = "Quantitative Coronary Analysis (QCA)"
        ws["A1"].font = header_font

        # Check if QCA results are available
        qca_results = None
        if hasattr(self, "qca_widget") and hasattr(self.qca_widget, "analysis_result"):
            qca_results = self.qca_widget.analysis_result
        elif hasattr(self, "sequential_qca_results"):
            qca_results = self.sequential_qca_results

        if not qca_results:
            ws["A3"] = "No QCA analysis results available"
            return

        # Check if we have valid sequential QCA results (dict with frame data)
        if isinstance(qca_results, dict) and not any(
            isinstance(k, int) and isinstance(v, dict) and v.get("success")
            for k, v in qca_results.items()
        ):
            ws["A3"] = "No valid QCA analysis results available"
            return

        # Column headers
        headers = [
            "Frame",
            "Cardiac Phase",
            "Stenosis (%)",
            "MLD (mm)",
            "Reference Diameter (mm)",
            "Lesion Length (mm)",
            "Area Stenosis (%)",
            "MLA (mm²)",
            "Reference Area (mm²)",
        ]

        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows
        row += 1

        # Handle different QCA result formats
        if isinstance(qca_results, dict) and "stenosis_info" in qca_results:
            # Single result format
            info = qca_results["stenosis_info"]
            ws.cell(row=row, column=1, value="Current")
            ws.cell(row=row, column=2, value="N/A")
            ws.cell(row=row, column=3, value=f"{info.get('percent_stenosis', 0):.1f}")
            ws.cell(row=row, column=4, value=f"{info.get('minimum_diameter', 0):.2f}")
            ws.cell(row=row, column=5, value=f"{info.get('reference_diameter', 0):.2f}")
            ws.cell(row=row, column=6, value=f"{info.get('lesion_length', 0):.2f}")
            ws.cell(row=row, column=7, value=f"{info.get('percent_area_stenosis', 0):.1f}")
            ws.cell(row=row, column=8, value=f"{info.get('mla', 0):.2f}")
            ws.cell(row=row, column=9, value=f"{info.get('reference_area', 0):.2f}")

        elif isinstance(qca_results, dict):
            # Sequential results format
            cardiac_phases = {
                "d2": "End-diastole",
                "s1": "Early-systole",
                "s2": "End-systole",
                "d1": "Mid-diastole",
            }

            for frame_idx in sorted(qca_results.keys()):
                if isinstance(frame_idx, int):
                    result = qca_results[frame_idx]
                    if result.get("success"):
                        ws.cell(row=row, column=1, value=frame_idx)

                        # Get cardiac phase if available
                        phase = ""
                        if hasattr(self, "ekg_parser") and self.ekg_parser:
                            frame_rate = (
                                self.dicom_parser.get_frame_rate()
                                if hasattr(self, "dicom_parser")
                                else 30.0
                            )
                            frame_time = frame_idx / frame_rate
                            phase_key = self.ekg_parser.get_phase_at_time(frame_time)
                            phase = cardiac_phases.get(phase_key, phase_key) if phase_key else ""

                        ws.cell(row=row, column=2, value=phase)
                        ws.cell(row=row, column=3, value=f"{result.get('percent_stenosis', 0):.1f}")
                        ws.cell(row=row, column=4, value=f"{result.get('mld', 0):.2f}")
                        ws.cell(
                            row=row, column=5, value=f"{result.get('reference_diameter', 0):.2f}"
                        )
                        ws.cell(row=row, column=6, value=f"{result.get('lesion_length', 0):.2f}")
                        ws.cell(
                            row=row, column=7, value=f"{result.get('percent_area_stenosis', 0):.1f}"
                        )
                        ws.cell(row=row, column=8, value=f"{result.get('mla', 0):.2f}")
                        ws.cell(row=row, column=9, value=f"{result.get('reference_area', 0):.2f}")
                        row += 1

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

    def _create_rws_sheet(self, wb):
        """Create RWS analysis data sheet"""
        from openpyxl.styles import Font

        ws = wb.create_sheet("RWS Analysis")

        # Header
        header_font = Font(bold=True, size=14)
        ws["A1"] = "Radial Wall Strain (RWS) Analysis"
        ws["A1"].font = header_font

        # Check if RWS results are available or can be calculated
        rws_results = None

        # First, check if Enhanced RWS results are available (preferred)
        if hasattr(self, "rws_enhanced_results") and self.rws_enhanced_results:
            rws_results = self.rws_enhanced_results
            logger.info("Using Enhanced RWS results for Excel export")
        elif hasattr(self, "rws_enhanced_widget") and hasattr(
            self.rws_enhanced_widget, "analysis_results"
        ):
            rws_results = self.rws_enhanced_widget.analysis_results
            logger.info("Using Enhanced RWS widget results for Excel export")
        elif hasattr(self, "rws_widget") and hasattr(self.rws_widget, "results"):
            rws_results = self.rws_widget.results
            logger.info("Using standard RWS widget results for Excel export")

        # If no widget results, try to calculate from QCA data if available
        if (
            not rws_results
            and hasattr(self, "sequential_qca_results")
            and self.sequential_qca_results
        ):
            try:
                rws_results = self._calculate_rws_from_qca_data(self.sequential_qca_results)
            except Exception as e:
                logger.warning(f"Failed to calculate RWS from QCA data: {e}")

        if not rws_results:
            ws["A3"] = "No RWS analysis results available"
            ws["A4"] = "RWS analysis results can be included by:"
            ws["A5"] = "1. Running RWS analysis from the QCA results menu, or"
            ws["A6"] = "2. Using the RWS Enhanced Analysis tab"
            return

        # RWS Summary
        row = 3
        ws[f"A{row}"] = "RWS Summary:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        # Extract RWS data from Enhanced or Standard analysis results
        if isinstance(rws_results, dict):
            # Check if this is Enhanced RWS results format
            if "stenosis_rws" in rws_results:
                # Enhanced RWS results format
                stenosis_rws = rws_results["stenosis_rws"]

                # RWS Percentage (Enhanced format)
                rws_val = stenosis_rws.get("rws_stenosis", 0)
                ws[f"A{row}"] = "Stenosis RWS Percentage:"
                ws[f"B{row}"] = f"{rws_val:.2f}%"
                row += 1

                # MLD Information from Enhanced analysis
                max_mld_phase = stenosis_rws.get("max_mld_phase", "unknown")
                min_mld_phase = stenosis_rws.get("min_mld_phase", "unknown")

                if "mld_values" in stenosis_rws:
                    mld_values = stenosis_rws["mld_values"]
                    if mld_values:
                        max_mld = max(mld_values.values()) if mld_values.values() else 0
                        min_mld = min(mld_values.values()) if mld_values.values() else 0

                        ws[f"A{row}"] = "Maximum MLD (mm):"
                        ws[f"B{row}"] = f"{max_mld:.2f}"
                        ws[f"C{row}"] = f"Phase: {max_mld_phase}"
                        row += 1

                        ws[f"A{row}"] = "Minimum MLD (mm):"
                        ws[f"B{row}"] = f"{min_mld:.2f}"
                        ws[f"C{row}"] = f"Phase: {min_mld_phase}"
                        row += 1

                        if max_mld > min_mld:
                            variation = max_mld - min_mld
                            ws[f"A{row}"] = "MLD Variation (mm):"
                            ws[f"B{row}"] = f"{variation:.2f}"
                            row += 1

                # Enhanced analysis method
                ws[f"A{row}"] = "Calculation Method:"
                ws[f"B{row}"] = "Enhanced RWS Analysis with outlier detection"
                row += 1

            else:
                # Standard RWS results format (fallback)
                # RWS Percentage
                if "rws_percentage" in rws_results:
                    rws_val = rws_results["rws_percentage"]
                    ws[f"A{row}"] = "RWS Percentage:"
                    ws[f"B{row}"] = f"{rws_val:.2f}%"
                    row += 1

                # Maximum MLD
                if "max_diameter" in rws_results:
                    ws[f"A{row}"] = "Maximum MLD (mm):"
                    ws[f"B{row}"] = f"{rws_results['max_diameter']:.2f}"
                    if "max_frame" in rws_results:
                        ws[f"C{row}"] = f"Frame {rws_results['max_frame']}"
                    row += 1

                # Minimum MLD
                if "min_diameter" in rws_results:
                    ws[f"A{row}"] = "Minimum MLD (mm):"
                    ws[f"B{row}"] = f"{rws_results['min_diameter']:.2f}"
                    if "min_frame" in rws_results:
                        ws[f"C{row}"] = f"Frame {rws_results['min_frame']}"
                    row += 1

                # MLD Variation
                if "max_diameter" in rws_results and "min_diameter" in rws_results:
                    variation = rws_results["max_diameter"] - rws_results["min_diameter"]
                    ws[f"A{row}"] = "MLD Variation (mm):"
                    ws[f"B{row}"] = f"{variation:.2f}"
                    row += 1

                # Calculation method
                if "method" in rws_results:
                    ws[f"A{row}"] = "Calculation Method:"
                    ws[f"B{row}"] = rws_results["method"]
                    row += 1

            # Common information for both formats
            # Frames analyzed
            frames_analyzed = rws_results.get("frames_analyzed") or len(
                rws_results.get("stenosis_rws", {}).get("mld_values", {})
            )
            if frames_analyzed:
                ws[f"A{row}"] = "Frames Analyzed:"
                ws[f"B{row}"] = str(frames_analyzed)
                row += 1

            # RWS Formula
            row += 1
            ws[f"A{row}"] = "RWS Formula:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = "Formula:"
            ws[f"B{row}"] = "(MLDmax - MLDmin) / MLDmin × 100%"
            row += 1

            # Clinical interpretation
            rws_val = rws_results.get("stenosis_rws", {}).get("rws_stenosis", 0) or rws_results.get(
                "rws_percentage", 0
            )
            ws[f"A{row}"] = "Clinical Interpretation:"
            if rws_val > 12:
                ws[f"B{row}"] = "⚠️ HIGH RWS (>12%) - Potential plaque vulnerability"
                ws[f"B{row}"].font = Font(color="FF0000")  # Red color
            else:
                ws[f"B{row}"] = "✓ NORMAL RWS (<12%) - Stable plaque characteristics"
                ws[f"B{row}"].font = Font(color="008000")  # Green color
            row += 1

        # Auto-fit columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15

    def _calculate_rws_from_qca_data(self, qca_results):
        """Calculate RWS using proper RWS analysis methodology"""
        if not qca_results or not isinstance(qca_results, dict):
            return None

        try:
            # Import the proper RWS analysis class
            from ..analysis.rws_analysis import RWSAnalysis

            # Create RWS analyzer
            rws_analyzer = RWSAnalysis()

            # Extract frame indices for the "beat" (all available frames)
            frame_indices = [k for k in qca_results.keys() if isinstance(k, int)]
            if len(frame_indices) < 2:
                return None

            # Use calibration factor if available
            calibration_factor = getattr(self, "calibration_factor", 0.12345) or 0.12345

            # Run RWS analysis
            rws_result = rws_analyzer.analyze_beat(
                qca_results=qca_results,
                beat_frames=frame_indices,
                calibration_factor=calibration_factor,
            )

            if not rws_result.get("success"):
                # Fallback to simple calculation if advanced analysis fails
                return self._simple_rws_calculation(qca_results)

            # Return results in expected format for Excel export
            return {
                "rws_percentage": rws_result.get("rws_at_mld", 0),
                "max_diameter": rws_result.get("mld_max_value", 0),
                "min_diameter": rws_result.get("mld_min_value", 0),
                "success": True,
                "method": "RWS Analysis with outlier detection",
                "frames_analyzed": rws_result.get("num_frames_analyzed", 0),
                "min_frame": (
                    rws_result.get("mld_min_frame", 0) + 1
                    if rws_result.get("mld_min_frame") is not None
                    else "N/A"
                ),
                "max_frame": (
                    rws_result.get("mld_max_frame", 0) + 1
                    if rws_result.get("mld_max_frame") is not None
                    else "N/A"
                ),
            }

        except Exception as e:
            logger.warning(f"Advanced RWS analysis failed: {e}, falling back to simple calculation")
            return self._simple_rws_calculation(qca_results)

    def _simple_rws_calculation(self, qca_results):
        """Fallback simple RWS calculation using correct formula"""
        if not qca_results or not isinstance(qca_results, dict):
            return None

        # Collect MLD values from QCA results
        mld_values = []
        mld_info_by_frame = {}

        for frame_idx, result in qca_results.items():
            if isinstance(frame_idx, int) and isinstance(result, dict) and result.get("success"):
                mld = result.get("mld", 0) or result.get("minimum_diameter", 0)
                if mld > 0:
                    mld_values.append(mld)
                    mld_info_by_frame[frame_idx] = mld

        if len(mld_values) < 2:
            return None

        # Find min and max MLD values
        max_mld = max(mld_values)
        min_mld = min(mld_values)

        if min_mld <= 0:
            return None

        # CORRECT RWS Formula: (MLDmax - MLDmin) / MLDmin × 100%
        rws_percentage = ((max_mld - min_mld) / min_mld) * 100

        # Find frames with min/max values
        min_frame = None
        max_frame = None
        for frame_idx, mld_val in mld_info_by_frame.items():
            if mld_val == min_mld and min_frame is None:
                min_frame = frame_idx + 1
            if mld_val == max_mld and max_frame is None:
                max_frame = frame_idx + 1

        return {
            "rws_percentage": rws_percentage,
            "max_diameter": max_mld,
            "min_diameter": min_mld,
            "success": True,
            "method": "Simple RWS calculation from QCA MLD values",
            "frames_analyzed": len(mld_values),
            "min_frame": min_frame,
            "max_frame": max_frame,
        }

    def _clean_filename(self, filename):
        """Clean filename by removing invalid characters for cross-platform compatibility"""
        import re

        # Remove or replace invalid characters for Windows/Mac/Linux filenames
        # Invalid characters: < > : " | ? * / \ and control characters
        invalid_chars = r'[<>:"|?*/\\]'
        cleaned = re.sub(invalid_chars, "_", filename)

        # Remove control characters and excessive whitespace
        cleaned = re.sub(r"[\x00-\x1f\x7f-\x9f]", "", cleaned)
        cleaned = re.sub(r"\s+", "_", cleaned)

        # Remove leading/trailing dots and spaces (Windows issue)
        cleaned = cleaned.strip(". ")

        # Limit length to reasonable size (Windows has 255 char limit)
        if len(cleaned) > 50:
            cleaned = cleaned[:50]

        # If empty after cleaning, return 'Patient'
        return cleaned if cleaned else "Patient"

    def _create_dicom_metadata_sheet(self, wb):
        """Create DICOM metadata sheet"""
        from openpyxl.styles import Font

        ws = wb.create_sheet("DICOM Metadata")

        # Header
        header_font = Font(bold=True, size=14)
        ws["A1"] = "DICOM Metadata"
        ws["A1"].font = header_font

        if (
            not hasattr(self, "dicom_parser")
            or not self.dicom_parser
            or not self.dicom_parser.dicom_data
        ):
            ws["A3"] = "No DICOM data available"
            return

        dicom_data = self.dicom_parser.dicom_data
        row = 3

        # Patient Information
        ws[f"A{row}"] = "Patient Information:"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        patient_fields = [
            ("PatientName", "Patient Name"),
            ("PatientID", "Patient ID"),
            ("PatientBirthDate", "Birth Date"),
            ("PatientSex", "Sex"),
            ("PatientAge", "Age"),
            ("PatientWeight", "Weight"),
            ("PatientSize", "Height"),
        ]

        for tag, label in patient_fields:
            if hasattr(dicom_data, tag):
                value = str(getattr(dicom_data, tag, ""))
                ws[f"A{row}"] = f"{label}:"
                ws[f"B{row}"] = value
                row += 1

        row += 1

        # Study Information
        ws[f"A{row}"] = "Study Information:"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        study_fields = [
            ("StudyDate", "Study Date"),
            ("StudyTime", "Study Time"),
            ("StudyID", "Study ID"),
            ("StudyDescription", "Study Description"),
            ("SeriesDescription", "Series Description"),
            ("Modality", "Modality"),
            ("Manufacturer", "Manufacturer"),
            ("ManufacturerModelName", "Model Name"),
            ("SoftwareVersions", "Software Version"),
        ]

        for tag, label in study_fields:
            if hasattr(dicom_data, tag):
                value = str(getattr(dicom_data, tag, ""))
                ws[f"A{row}"] = f"{label}:"
                ws[f"B{row}"] = value
                row += 1

        row += 1

        # Image Information
        ws[f"A{row}"] = "Image Information:"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Image dimensions
        if hasattr(self.dicom_parser, "pixel_array") and self.dicom_parser.pixel_array is not None:
            shape = self.dicom_parser.pixel_array.shape
            ws[f"A{row}"] = "Image Dimensions:"
            if len(shape) == 3:
                ws[f"B{row}"] = f"{shape[2]} x {shape[1]} pixels, {shape[0]} frames"
            else:
                ws[f"B{row}"] = f"{shape[1]} x {shape[0]} pixels"
            row += 1

        image_fields = [
            ("PixelSpacing", "Pixel Spacing"),
            ("ImagerPixelSpacing", "Imager Pixel Spacing"),
            ("SliceThickness", "Slice Thickness"),
            ("KVP", "kVp"),
            ("ExposureTime", "Exposure Time"),
            ("XRayTubeCurrent", "Tube Current"),
            ("FilterType", "Filter Type"),
            ("FrameRate", "Frame Rate"),
            ("CineRate", "Cine Rate"),
        ]

        for tag, label in image_fields:
            if hasattr(dicom_data, tag):
                value = str(getattr(dicom_data, tag, ""))
                ws[f"A{row}"] = f"{label}:"
                ws[f"B{row}"] = value
                row += 1

        # Auto-fit columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
